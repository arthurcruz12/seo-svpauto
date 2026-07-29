import os
import tempfile
import unittest
import time
import shutil
from io import BytesIO
from unittest.mock import patch

from openpyxl import Workbook, load_workbook


_database = tempfile.NamedTemporaryFile(suffix=".sqlite3", delete=False)
_database.close()
os.environ["SEO_DATABASE_PATH"] = _database.name
_file_storage = tempfile.mkdtemp(prefix="seo-files-")
os.environ["SEO_FILE_STORAGE_PATH"] = _file_storage

from backend.app.main import (  # noqa: E402
    build_issues,
    build_document_intelligence,
    build_simple_pdf,
    ai_analyze,
    normalize,
    parse_number,
    read_rows,
    row_to_debt,
    row_to_inventory,
    rows_to_dataset,
)
from backend.app.schemas import AiQuestionRequest  # noqa: E402
from backend.app.security import authenticate_user  # noqa: E402
from backend.app.store import (  # noqa: E402
    connect,
    consume_rate_limit,
    get_mfa_challenge,
    get_uploaded_file,
    increment_mfa_attempt,
    list_uploaded_files,
    save_mfa_challenge,
    save_uploaded_file,
)
from backend.app.svp_billing import transform_svp_billing_workbook  # noqa: E402


class ImportRegressionTests(unittest.TestCase):
    @classmethod
    def tearDownClass(cls):
        try:
            os.unlink(_database.name)
        except (FileNotFoundError, PermissionError):
            pass
        shutil.rmtree(_file_storage, ignore_errors=True)

    def test_due_date_is_not_converted_to_debt_amount(self):
        row = {"entidade": "Cliente X", "vencimento": "2026-07-31"}
        normalized = {normalize(key): value for key, value in row.items()}
        self.assertIsNone(row_to_debt(row, normalized, 0))

    def test_payment_keeps_invoice_issue_and_due_dates(self):
        row = {
            "fatura": "FT CUSA/78",
            "cliente": "Cliente Coimbra",
            "valor em aberto": "82,00",
            "data emissão": "2026-07-21",
            "data vencimento": "2026-08-20",
            "prazo": 30,
        }
        normalized = {normalize(key): value for key, value in row.items()}
        debt = row_to_debt(row, normalized, 0)
        self.assertEqual(debt["invoice"], "FT CUSA/78")
        self.assertEqual(debt["issueDate"], "2026-07-21")
        self.assertEqual(debt["dueDate"], "2026-08-20")
        self.assertEqual(debt["dueDays"], 30)

    def test_svp_billing_is_split_grouped_and_recalculated(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["ID", "Documento", "Data Doc.", "Entidade", "Total", "Total liquido", "Total IVA", "Estado", "Doc. Fornecedor", "Nº Enc.", "Canal", "Vendedor", "F. Liquidação"])
        sheet.append([3, "FR PUSA/2", "2026-07-21", "Cliente P", 123, 100, 23, "Liquidado", 21, 21, "Balcão", "Vendedor", "MB"])
        sheet.append([1, "NC CUSA/1", "2026-07-21", "Cliente C", 61.5, 50, 11.5, "Liquidado", 21, 21, "Balcão", "Vendedor", "MB"])
        sheet.append([2, "FR CNOV/1", "2026-07-21", "Anulado", 20, 16.26, 3.74, "Anulado", 21, 21, "Balcão", "Vendedor", "MB"])
        sheet.append([4, "GT PUSA/1", "2026-07-21", "Guia", 0, 0, 0, "Ativo", 21, 21, "Balcão", "Vendedor", "NU"])
        buffer = BytesIO()
        workbook.save(buffer)

        transformed, summary = transform_svp_billing_workbook(buffer.getvalue(), "faturação 21.07.xlsx")
        result = load_workbook(BytesIO(transformed), data_only=False)
        rows = list(result.active.iter_rows(values_only=True))
        result.close()

        self.assertEqual(summary["includedRows"], 2)
        self.assertEqual(summary["excludedCancelled"], 1)
        self.assertEqual(summary["excludedNonBilling"], 1)
        self.assertEqual(summary["totalAmount"], 61.5)
        self.assertIn("COIMBRA", [row[0] for row in rows])
        self.assertIn("PICOTO", [row[0] for row in rows])
        credit_row = next(row for row in rows if row[1] == "NC CUSA/1")
        self.assertEqual(credit_row[4:7], (-61.5, -50, -11.5))
        self.assertTrue(any(row[3] == "TOTAL GERAL" and str(row[4]).startswith("=SUM(") for row in rows))

    def test_zero_stock_is_kept_as_critical_inventory(self):
        row = {"produto": "Farol", "stock": 0, "referencia": "SKU-0"}
        normalized = {normalize(key): value for key, value in row.items()}
        item = row_to_inventory(row, normalized, 0)
        self.assertIsNotNone(item)
        self.assertEqual(item["stock"], 0)
        self.assertEqual(item["alert"], "Stock crítico")

    def test_svp_inventory_rule_and_physical_difference(self):
        row = {
            "produto": "Motor usado",
            "referencia": "MOT-1",
            "codigo unidade": "  cusa ",
            "stock": 10,
            "quantidade sistema": 10,
            "quantidade fisica": 8,
            "custo unitario": "1.250,50",
            "localizacao": "A-03",
        }
        normalized = {normalize(key): value for key, value in row.items()}
        item = row_to_inventory(row, normalized, 0)
        self.assertEqual(item["unit"], "Coimbra")
        self.assertEqual(item["stockType"], "Usado")
        self.assertEqual(item["differenceQuantity"], -2)
        self.assertEqual(item["stockValue"], 10004.0)
        self.assertEqual(item["movementType"], "Existente")
        self.assertEqual(item["movementQuantity"], 8)
        self.assertEqual(item["confidence"], 100)

    def test_inventory_separates_purchase_sale_and_scrap_movements(self):
        cases = [
            ("Compra de motor novo", "CNOV", "Compra", 3),
            ("Venda de caixa usada", "PUSA", "Venda", -2),
            ("Saida para sucata", "CUSA", "Sucata", -4),
        ]
        for index, (description, code, movement, quantity) in enumerate(cases):
            row = {"produto": description, "codigo unidade": code, "stock": 10, "quantidade sistema": 10, "quantidade fisica": 9, "quantidade movimento": abs(quantity)}
            normalized = {normalize(key): value for key, value in row.items()}
            item = row_to_inventory(row, normalized, index)
            self.assertEqual(item["movementType"], movement)
            self.assertEqual(item["movementQuantity"], quantity)
        self.assertIn("Divergência", item["alert"])

    def test_unknown_inventory_code_requires_review(self):
        row = {"produto": "Peça sem regra", "referencia": "X-1", "unidade": "XYZ", "stock": 2}
        normalized = {normalize(key): value for key, value in row.items()}
        item = row_to_inventory(row, normalized, 0)
        self.assertEqual(item["unit"], "Não identificado")
        self.assertEqual(item["validationState"], "Revisão necessária")
        self.assertEqual(item["confidence"], 45)

    def test_anomalies_cover_documents_inventory_and_overdue_payments(self):
        document_intelligence = {
            "documents": [
                {
                    "number": "FT-1",
                    "entity": "Fornecedor",
                    "totalAmount": 123,
                    "validations": ["Possível documento duplicado"],
                }
            ]
        }
        inventory = [
            {
                "ref": "SKU-1",
                "warehouse": "Coimbra",
                "unit": "Coimbra",
                "differenceQuantity": -2,
                "systemQuantity": 10,
                "physicalQuantity": 8,
                "unitCost": 50,
                "stockValue": 400,
                "confidence": 100,
            }
        ]
        debts = [{"invoice": "FT-2", "entity": "Cliente", "amount": 900, "dueDays": 45, "state": "Em atraso"}]

        issues = build_issues([], document_intelligence, inventory, debts)

        self.assertEqual(len(issues), 3)
        self.assertTrue(any("duplicado" in issue["issue"].lower() for issue in issues))
        self.assertTrue(any("inventário" in issue["issue"].lower() for issue in issues))
        self.assertTrue(any("vencido" in issue["issue"].lower() for issue in issues))
        self.assertTrue(all(issue["status"] == "Alerta" for issue in issues))

    def test_saft_invoice_reads_nested_totals(self):
        xml = """<?xml version='1.0' encoding='UTF-8'?>
        <AuditFile xmlns='urn:OECD:StandardAuditFile-Tax:PT_1.04_01'>
          <SourceDocuments><SalesInvoices><Invoice>
            <InvoiceNo>FT 1/1</InvoiceNo><InvoiceDate>2026-07-01</InvoiceDate>
            <CustomerID>C1</CustomerID><Line><Description>Venda de peça</Description><CreditAmount>100.00</CreditAmount></Line>
            <DocumentTotals><TaxPayable>23.00</TaxPayable><NetTotal>100.00</NetTotal><GrossTotal>123.00</GrossTotal></DocumentTotals>
          </Invoice></SalesInvoices></SourceDocuments>
        </AuditFile>""".encode("utf-8")
        rows = read_rows(xml, "xml")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["GrossTotal"], "123.00")
        dataset = rows_to_dataset("saft.xml", rows)
        self.assertEqual(dataset["documentIntelligence"]["totals"]["total"], 123.0)

    def test_credit_notes_are_always_negative(self):
        analysis = build_document_intelligence(
            "notas.csv",
            [{"documento": "NC 1", "descricao": "Nota de crédito", "valor sem iva": "100", "iva": "23", "total": "123"}],
        )
        self.assertEqual(analysis["totals"]["total"], -123.0)

    def test_text_pdf_invoice_is_extracted(self):
        pdf = build_simple_pdf(["Fatura: FT-2026-9", "Data: 2026-07-15", "Fornecedor: Peças Demo", "Subtotal: 100.00", "IVA: 23.00", "Total: 123.00"])
        rows = read_rows(pdf, "pdf")
        self.assertEqual(rows[0]["documento"], "FT-2026-9")
        self.assertEqual(rows[0]["total"], "123.00")

    def test_excel_footer_total_is_not_imported_as_document(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["ID", "Documento", "Total"])
        sheet.append([1, "FT CUSA/1", 123])
        sheet.append(["Total", None, 123])
        output = BytesIO()
        workbook.save(output)
        rows = read_rows(output.getvalue(), "xlsx")
        self.assertEqual(rows, [{"ID": 1, "Documento": "FT CUSA/1", "Total": 123}])

    @patch("backend.app.main.extract_image_text")
    def test_local_ocr_image_invoice_is_extracted(self, extract_image_text):
        extract_image_text.return_value = "Fatura: FT-LOCAL-1\nData: 21/07/2026\nFornecedor: Demo\nSubtotal: 100,00\nIVA: 23,00\nTotal: 123,00"
        rows = read_rows(b"local-image", "jpg")
        self.assertEqual(rows[0]["documento"], "FT-LOCAL-1")
        self.assertEqual(rows[0]["total"], "123,00")

    @patch("backend.app.main.extract_scanned_pdf_text")
    @patch("backend.app.main.PdfReader")
    def test_local_ocr_scanned_pdf_invoice_is_extracted(self, pdf_reader, extract_scanned_pdf_text):
        pdf_reader.return_value.pages = []
        extract_scanned_pdf_text.return_value = "Fatura: FT-SCAN-1\nData: 21/07/2026\nTotal: 250,00"
        rows = read_rows(b"scanned-pdf", "pdf")
        self.assertEqual(rows[0]["documento"], "FT-SCAN-1")
        self.assertEqual(rows[0]["total"], "250,00")

    def test_original_upload_is_persisted_with_integrity_hash(self):
        content = b"invoice-original-content"
        stored = save_uploaded_file("company-test", "admin@example.com", "fatura.pdf", "application/pdf", "documentos", content)
        self.assertEqual(stored["sizeBytes"], len(content))
        self.assertEqual(len(stored["sha256"]), 64)
        self.assertEqual(list_uploaded_files("company-test")[0]["filename"], "fatura.pdf")
        self.assertEqual(get_uploaded_file("company-test", stored["id"])["content"], content)
        self.assertEqual(list_uploaded_files("other-company"), [])

    def test_ai_explainability_text_keeps_portuguese_encoding(self):
        user = authenticate_user("admin@seo.local", "Seo-Admin-2026")
        result = ai_analyze(AiQuestionRequest(question="Qual é o risco atual?"), user)
        explanation = " ".join(result["explainability"]["dataSources"]) + result["explainability"]["method"]
        self.assertIn("inventário", explanation)
        self.assertIn("conciliação", explanation)
        self.assertNotIn("?", explanation)

    def test_ai_analysis_levels_change_fallback_detail(self):
        user = authenticate_user("admin@seo.local", "Seo-Admin-2026")
        quick = ai_analyze(AiQuestionRequest(question="Qual é o risco atual?", analysis_level="Rápido"), user)
        audit = ai_analyze(AiQuestionRequest(question="Qual é o risco atual?", analysis_level="Auditoria"), user)
        self.assertEqual(quick["analysisLevel"], "Rápido")
        self.assertEqual(audit["analysisLevel"], "Auditoria")
        self.assertLess(len(quick["answer"]), len(audit["answer"]))
        self.assertIn("Para auditoria", audit["answer"])

    def test_misaligned_excel_finds_headers_and_recalculates_total(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Relatório de faturas — julho 2026"])
        sheet.append([None, None, None])
        sheet.append(["Documento", "Descrição", "Fornecedor", "Valor sem IVA", "IVA", "Total"])
        sheet.append(["FT 1", "Fatura fornecedor de peças", "Fornecedor X", 100, 23, None])
        sheet.append(["TOTAL GERAL", None, None, 100, 23, 123])
        buffer = BytesIO()
        workbook.save(buffer)

        rows = read_rows(buffer.getvalue(), "xlsx")
        self.assertEqual(len(rows), 1)
        dataset = rows_to_dataset("faturas-desconfiguradas.xlsx", rows)
        document = dataset["documentIntelligence"]["documents"][0]
        self.assertEqual(document["netAmount"], 100)
        self.assertEqual(document["vatAmount"], 23)
        self.assertEqual(document["totalAmount"], 123)
        self.assertFalse(document["needsReview"])
        self.assertIn("Total recalculado", document["validations"][0])

    def test_disordered_columns_and_portuguese_numbers(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Exportação manual", None, "Não editar"])
        sheet.append([None, "IVA", "Entidade", "Total", "Documento", "Valor sem IVA", "Descrição"])
        sheet.append([None, "234,56", "Fornecedor PT", "1.254,56", "FT 22", "1.020,00", "Fatura de mercadoria"])
        output = BytesIO()
        workbook.save(output)

        dataset = rows_to_dataset("colunas-trocadas.xlsx", read_rows(output.getvalue(), "xlsx"))
        document = dataset["documentIntelligence"]["documents"][0]
        self.assertEqual(document["netAmount"], 1020)
        self.assertEqual(document["vatAmount"], 234.56)
        self.assertEqual(document["totalAmount"], 1254.56)

    def test_missing_net_and_vat_are_calculated_from_rate(self):
        workbook = Workbook()
        sheet = workbook.active
        for _ in range(7):
            sheet.append(["Linha de apresentação sem dados"])
        sheet.append(["Documento", "Descrição", "Entidade", "Total", "Taxa IVA"])
        sheet.append(["FT 30", "Fatura fornecedor", "Fornecedor Taxa", 123, 23])
        output = BytesIO()
        workbook.save(output)

        dataset = rows_to_dataset("apenas-total-taxa.xlsx", read_rows(output.getvalue(), "xlsx"))
        document = dataset["documentIntelligence"]["documents"][0]
        self.assertEqual(document["netAmount"], 100)
        self.assertEqual(document["vatAmount"], 23)
        self.assertEqual(document["totalAmount"], 123)
        self.assertEqual(dataset["documentIntelligence"]["stats"]["corrected"], 1)

    def test_positive_credit_note_is_normalized_and_footer_ignored(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["NOTAS DE CRÉDITO"])
        sheet.append(["Documento", "Descrição", "Entidade", "Valor sem IVA", "IVA", "Total"])
        sheet.append(["NC 7", "Nota de crédito devolução", "Fornecedor X", 50, 11.5, 61.5])
        sheet.append(["TOTAIS", None, None, 50, 11.5, 61.5])
        output = BytesIO()
        workbook.save(output)

        rows = read_rows(output.getvalue(), "xlsx")
        self.assertEqual(len(rows), 1)
        document = rows_to_dataset("notas-desconfiguradas.xlsx", rows)["documentIntelligence"]["documents"][0]
        self.assertEqual(document["documentType"], "Nota de crédito")
        self.assertEqual(document["totalAmount"], -61.5)

    def test_duplicate_and_blank_headers_do_not_drop_the_row(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Documento", "Descrição", "Entidade", "IVA", "IVA", None, "Total"])
        sheet.append(["FT 40", "Fatura serviços", "Fornecedor Y", 23, None, "ignorar", 123])
        output = BytesIO()
        workbook.save(output)

        rows = read_rows(output.getvalue(), "xlsx")
        self.assertEqual(len(rows), 1)
        self.assertIn("coluna_6", rows[0])
        document = rows_to_dataset("duplicados.xlsx", rows)["documentIntelligence"]["documents"][0]
        self.assertEqual(document["vatAmount"], 23)
        self.assertEqual(document["totalAmount"], 123)

    def test_large_excel_processes_all_totals_and_limits_visual_payload(self):
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet("Faturas")
        sheet.append(["Documento", "Descrição", "Entidade", "Valor sem IVA", "IVA", "Total"])
        row_count = 20_000
        for index in range(row_count):
            sheet.append([f"FT {index + 1}", "Venda loja online", "Cliente", 10, 2.3, 12.3])
        output = BytesIO()
        workbook.save(output)

        started = time.perf_counter()
        rows = read_rows(output.getvalue(), "xlsx")
        dataset = rows_to_dataset("faturas-grandes.xlsx", rows)
        elapsed = time.perf_counter() - started

        self.assertEqual(dataset["summary"]["rowsRead"], row_count)
        self.assertEqual(dataset["documentIntelligence"]["stats"]["processed"], row_count)
        self.assertEqual(dataset["documentIntelligence"]["totals"]["total"], 246_000)
        self.assertEqual(len(dataset["documentIntelligence"]["documents"]), 1000)
        self.assertEqual(len(dataset["classifiedMovements"]), 1000)
        self.assertLess(elapsed, 30)

    def test_number_parser_handles_common_accounting_formats(self):
        cases = {
            "1.234,56 €": 1234.56,
            "1,234.56": 1234.56,
            "(450,25)": -450.25,
            "- 99,90": -99.9,
            "2.500": 2500,
            "": 0,
        }
        for raw, expected in cases.items():
            with self.subTest(raw=raw):
                self.assertEqual(parse_number(raw), expected)

    def test_invalid_date_and_inconsistent_total_require_review(self):
        analysis = build_document_intelligence(
            "erros.xlsx",
            [{"documento": "FT ERR", "data": "31/99/2026", "entidade": "Fornecedor", "valor sem iva": 100, "iva": 23, "total": 150}],
        )
        document = analysis["documents"][0]
        self.assertTrue(document["needsReview"])
        self.assertTrue(any("Data inválida" in item for item in document["validations"]))
        self.assertTrue(any("não corresponde" in item for item in document["validations"]))

    def test_duplicate_documents_are_detected(self):
        row = {"documento": "FT DUP", "entidade": "Fornecedor", "total": 123}
        analysis = build_document_intelligence("duplicados.xlsx", [row, row.copy()])
        self.assertEqual(analysis["stats"]["duplicates"], 1)
        self.assertTrue(analysis["documents"][1]["needsReview"])

    def test_multiple_excel_sheets_are_combined(self):
        workbook = Workbook()
        first = workbook.active
        first.title = "Janeiro"
        second = workbook.create_sheet("Fevereiro")
        for sheet, number in ((first, "FT JAN"), (second, "FT FEV")):
            sheet.append(["Documento", "Descrição", "Entidade", "Total"])
            sheet.append([number, "Venda loja", "Cliente", 100])
        output = BytesIO()
        workbook.save(output)
        rows = read_rows(output.getvalue(), "xlsx")
        self.assertEqual(len(rows), 2)
        self.assertEqual({row["Documento"] for row in rows}, {"FT JAN", "FT FEV"})

    def test_csv_semicolon_and_tab_delimiters(self):
        semicolon = read_rows("documento;descricao;entidade;total\nFT1;Venda;Cliente;123\n".encode(), "csv")
        tabbed = read_rows("documento\tdescricao\tentidade\ttotal\nFT2\tVenda\tCliente\t50\n".encode(), "txt")
        self.assertEqual(semicolon[0]["documento"], "FT1")
        self.assertEqual(tabbed[0]["documento"], "FT2")

    def test_malformed_xml_is_rejected_cleanly(self):
        from fastapi import HTTPException

        with self.assertRaises(HTTPException) as error:
            read_rows(b"<AuditFile><Invoice>", "xml")
        self.assertEqual(error.exception.status_code, 400)

    def test_uploaded_file_content_is_kept_outside_sqlite(self):
        content = b"large-binary-document" * 100
        stored = save_uploaded_file("company-external", "owner@example.com", "fatura.pdf", "application/pdf", "documentos", content)
        with connect() as connection:
            row = connection.execute(
                "SELECT length(content) AS blob_size, storage_path FROM uploaded_files WHERE id = ?",
                (stored["id"],),
            ).fetchone()
        self.assertEqual(row["blob_size"], 0)
        self.assertTrue(row["storage_path"])
        self.assertEqual(get_uploaded_file("company-external", stored["id"])["content"], content)

    def test_security_state_is_persistent_and_rate_limited(self):
        save_mfa_challenge("persistent-challenge", "owner@example.com", "hash", "2999-01-01T00:00:00+00:00")
        self.assertEqual(get_mfa_challenge("persistent-challenge")["email"], "owner@example.com")
        self.assertEqual(increment_mfa_attempt("persistent-challenge"), 1)
        self.assertTrue(consume_rate_limit("persistent-key", 2, "2000-01-01T00:00:00+00:00"))
        self.assertTrue(consume_rate_limit("persistent-key", 2, "2000-01-01T00:00:00+00:00"))
        self.assertFalse(consume_rate_limit("persistent-key", 2, "2000-01-01T00:00:00+00:00"))


if __name__ == "__main__":
    unittest.main()
