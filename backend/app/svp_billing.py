from __future__ import annotations

from datetime import datetime
from io import BytesIO
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DOCUMENT_RE = re.compile(r"^(FR|FT|NC)\s+(CNOV|CUSA|PNOV|PUSA|POFI)/", re.IGNORECASE)
LOCATION_SERIES = {
    "COIMBRA": ("CNOV", "CUSA"),
    "PICOTO": ("PNOV", "POFI", "PUSA"),
}
DOCUMENT_TYPES = ("FR", "FT", "NC")


def is_svp_billing_workbook(content: bytes) -> bool:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        found = 0
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(min_row=2, max_col=2, values_only=True):
                if DOCUMENT_RE.match(str(row[1] or "").strip()):
                    found += 1
                    if found >= 2:
                        workbook.close()
                        return True
        workbook.close()
    except Exception:
        return False
    return False


def transform_svp_billing_workbook(content: bytes, source_name: str) -> tuple[bytes, dict]:
    source = load_workbook(BytesIO(content), data_only=True)
    source_sheet = source.active
    source_row_count = max(source_sheet.max_row - 2, 0)
    headers = [cell.value for cell in source_sheet[1]][:13]
    records: list[dict] = []
    excluded_cancelled = 0
    excluded_non_billing = 0

    for row in source_sheet.iter_rows(min_row=2, max_col=13, values_only=True):
        document = str(row[1] or "").strip()
        match = DOCUMENT_RE.match(document)
        if not match:
            first_cell = str(row[0] or "").strip().casefold()
            if first_cell not in {"total", "total geral"} and any(value not in (None, "") for value in row):
                excluded_non_billing += 1
            continue
        if str(row[7] or "").strip().casefold() == "anulado":
            excluded_cancelled += 1
            continue
        document_type, series = match.groups()
        values = list(row)
        values[2] = _to_excel_date(values[2])
        for index in (4, 5, 6):
            amount = _to_number(values[index])
            values[index] = -abs(amount) if document_type.upper() == "NC" else abs(amount)
        records.append({
            "values": values,
            "type": document_type.upper(),
            "series": series.upper(),
            "sort_id": _sort_id(values[0]),
        })
    source.close()

    output = Workbook()
    sheet = output.active
    sheet.title = _sheet_title(source_name)
    sheet.append(headers)
    _style_header(sheet, 1)
    subtotal_rows: list[int] = []
    group_counts: dict[str, int] = {}

    for location, series_order in LOCATION_SERIES.items():
        section_records = [record for record in records if record["series"] in series_order]
        if not section_records:
            continue
        sheet.append([location] + [None] * 12)
        _style_location(sheet, sheet.max_row)
        for document_type in DOCUMENT_TYPES:
            for series in series_order:
                group = sorted(
                    [record for record in section_records if record["type"] == document_type and record["series"] == series],
                    key=lambda record: record["sort_id"],
                )
                if not group:
                    continue
                start_row = sheet.max_row + 1
                for record in group:
                    sheet.append(record["values"])
                    _style_data_row(sheet, sheet.max_row)
                end_row = sheet.max_row
                label = f"{document_type} {series}"
                sheet.append([None, None, None, label] + [None] * 9)
                subtotal_row = sheet.max_row
                for column in (5, 6, 7):
                    letter = get_column_letter(column)
                    sheet.cell(subtotal_row, column, f"=SUM({letter}{start_row}:{letter}{end_row})")
                _style_subtotal(sheet, subtotal_row)
                subtotal_rows.append(subtotal_row)
                group_counts[label] = len(group)
        sheet.append([None] * 13)

    sheet.append([None, None, None, "TOTAL GERAL"] + [None] * 9)
    total_row = sheet.max_row
    for column in (5, 6, 7):
        letter = get_column_letter(column)
        refs = ",".join(f"{letter}{row}" for row in subtotal_rows)
        sheet.cell(total_row, column, f"=SUM({refs})" if refs else "=0")
    _style_total(sheet, total_row)
    _set_dimensions(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:M{total_row}"

    output_bytes = BytesIO()
    output.save(output_bytes)
    output.close()
    return output_bytes.getvalue(), {
        "sourceRows": source_row_count,
        "includedRows": len(records),
        "excludedCancelled": excluded_cancelled,
        "excludedNonBilling": excluded_non_billing,
        "groups": group_counts,
        "totalRow": total_row,
        "totalAmount": round(sum(float(record["values"][4] or 0) for record in records), 2),
    }


def _sheet_title(source_name: str) -> str:
    stem = source_name.rsplit(".", 1)[0]
    clean = re.sub(r"[^A-Za-zÀ-ÿ0-9 ._-]", "", stem).strip() or "Faturação organizada"
    return clean[:31]


def _sort_id(value) -> tuple[int, str]:
    text = str(value or "")
    match = re.search(r"\d+", text)
    return (int(match.group()) if match else 10**12, text)


def _to_number(value) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "0").strip().replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _to_excel_date(value):
    if isinstance(value, datetime):
        return value
    text = str(value or "").strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    return value


def _style_header(sheet, row: int) -> None:
    for cell in sheet[row]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[row].height = 24


def _style_location(sheet, row: int) -> None:
    side = Side(style="medium", color="7F7F7F")
    for cell in sheet[row]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(name="Calibri", size=11, bold=True, color="1F1F1F")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=side, bottom=side)
    sheet.row_dimensions[row].height = 22


def _style_data_row(sheet, row: int) -> None:
    for column in range(1, 14):
        cell = sheet.cell(row, column)
        cell.font = Font(name="Calibri", size=11)
        cell.alignment = Alignment(vertical="top", wrap_text=column in (4, 11, 12))
    sheet.cell(row, 3).number_format = "dd/mm/yyyy"
    for column in (5, 6, 7):
        sheet.cell(row, column).number_format = "#,##0.00;[Red]-#,##0.00"


def _style_subtotal(sheet, row: int) -> None:
    thin = Side(style="thin", color="A6A6A6")
    for column in range(4, 8):
        cell = sheet.cell(row, column)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill("solid", fgColor="E2F0D9")
        cell.border = Border(top=thin, bottom=thin)
    for column in (5, 6, 7):
        sheet.cell(row, column).number_format = "#,##0.00;[Red]-#,##0.00"


def _style_total(sheet, row: int) -> None:
    medium = Side(style="medium", color="1F4E78")
    for column in range(4, 8):
        cell = sheet.cell(row, column)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.border = Border(top=medium, bottom=medium)
    for column in (5, 6, 7):
        sheet.cell(row, column).number_format = "#,##0.00;[Red]-#,##0.00"


def _set_dimensions(sheet) -> None:
    widths = [12, 18, 13, 38, 14, 15, 13, 14, 18, 20, 29, 31, 16]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in range(2, sheet.max_row + 1):
        if sheet.row_dimensions[row].height is None:
            sheet.row_dimensions[row].height = 20
