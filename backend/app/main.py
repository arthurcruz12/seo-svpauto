from datetime import datetime, timezone
from io import BytesIO
import csv
import hmac
import hashlib
import json
import logging
import math
import os
import re
import unicodedata
import xml.etree.ElementTree as ET
from itertools import chain
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from fastapi import Depends, FastAPI, File, HTTPException, Request as FastAPIRequest, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from pypdf import PdfReader
from starlette.concurrency import run_in_threadpool

from .audit import list_audit_events, record_audit_event
from .ai_provider import generate_contextual_answer
from .local_ocr import extract_image_text, extract_scanned_pdf_text
from .services.ocr_service import OcrExecutionError, OcrValidationError, read_full_file
from .schemas import (
    AccountProfile,
    AiQuestionRequest,
    AuditEvent,
    BillingCheckoutRequest,
    ChallengeResponse,
    LoginRequest,
    MfaRequest,
    RegisterRequest,
    SnapshotCreateRequest,
    StrategyFitRequest,
    TokenResponse,
)
from .security import (
    MFA_TTL_SECONDS,
    account_profile,
    authenticate_user,
    create_access_token,
    create_mfa_challenge,
    enforce_rate_limit,
    get_current_user,
    register_client,
    require_permission,
    verify_mfa_challenge,
)
from .store import (
    append_reconciliation_issues,
    add_payment_event,
    get_subscription,
    get_company,
    get_operational_state,
    get_uploaded_file,
    get_ai_messages,
    list_debt_items,
    list_inventory_items,
    list_metric_snapshots,
    list_uploaded_files,
    list_reconciliation_issues,
    mark_debt_paid,
    register_inventory_sale,
    replace_operational_dataset,
    resolve_all_reconciliation_issues,
    resolve_reconciliation_issue,
    save_metric_snapshot,
    save_operational_state,
    save_uploaded_file,
    save_ai_exchange,
    upsert_subscription,
)
from .strategy import score_strategy_fit, strategy_operating_model
from .svp_billing import is_svp_billing_workbook, transform_svp_billing_workbook


app = FastAPI(title="SEO API", version="0.3.0")
logger = logging.getLogger(__name__)

APP_ENV = os.getenv("SEO_ENV", "development").lower()
FRONTEND_URL = os.getenv("SEO_FRONTEND_URL", "http://127.0.0.1:5173")
STRIPE_SECRET_KEY = os.getenv("STRIPE_SECRET_KEY", "")
STRIPE_WEBHOOK_SECRET = os.getenv("STRIPE_WEBHOOK_SECRET", "")
STRIPE_PRICE_STARTER = os.getenv("STRIPE_PRICE_STARTER", "")
STRIPE_PRICE_PROFESSIONAL = os.getenv("STRIPE_PRICE_PROFESSIONAL", "")
STRIPE_PRICE_BUSINESS = os.getenv("STRIPE_PRICE_BUSINESS", "")
try:
    MAX_UPLOAD_BYTES = max(1024 * 1024, int(os.getenv("SEO_MAX_UPLOAD_MB", "50")) * 1024 * 1024)
except ValueError:
    MAX_UPLOAD_BYTES = 50 * 1024 * 1024
MAX_OCR_FILE_SIZE = 25 * 1024 * 1024
STRIPE_PRICES = {
    "starter": STRIPE_PRICE_STARTER,
    "professional": STRIPE_PRICE_PROFESSIONAL,
    "business": STRIPE_PRICE_BUSINESS,
}
DEFAULT_CORS_ORIGINS = ",".join(
    [
        FRONTEND_URL,
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost:4173",
        "http://127.0.0.1:4173",
    ]
)
CORS_ORIGINS = [origin.strip() for origin in os.getenv("SEO_CORS_ORIGINS", DEFAULT_CORS_ORIGINS).split(",") if origin.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST"],
    allow_headers=["Authorization", "Content-Type"],
)


@app.middleware("http")
async def security_headers(request: FastAPIRequest, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    if APP_ENV == "production":
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return response

ACCOUNT_RULES = [
    ("11", "Caixa", ["caixa", "numerario", "dinheiro", "fundo de caixa"], "Débito"),
    ("12", "Depósitos à ordem", ["transferencia", "banco", "mb way", "terminal", "tpa", "deposito"], "Débito"),
    ("21", "Clientes", ["cliente", "recebimento", "fatura cliente", "recibo"], "Débito"),
    ("22", "Fornecedores", ["fornecedor", "pagamento fornecedor", "fatura fornecedor", "compra fornecedor"], "Crédito"),
    ("24", "Estado e outros entes públicos", ["iva", "seguranca social", "at", "financas", "imposto", "retencao", "irc"], "Crédito"),
    ("31", "Compras", ["compra", "mercadoria", "pecas", "stock", "inventario", "sucata"], "Débito"),
    ("32", "Mercadorias", ["mercadorias", "existencias", "armazem"], "Débito"),
    ("61", "Custo das mercadorias vendidas e das matérias consumidas", ["cmv", "cmvmc", "custo mercadoria", "custo das vendas"], "Débito"),
    ("62", "Fornecimentos e serviços externos", ["renda", "energia", "agua", "internet", "telefone", "seguro", "transporte", "combustivel", "software", "comissao", "ovoko", "recambio", "shopify"], "Débito"),
    ("63", "Gastos com o pessoal", ["salario", "ordenado", "vencimento", "subsidio", "funcionario"], "Débito"),
    ("68", "Outros gastos e perdas", ["multa", "coima", "juros", "perda", "regularizacao", "imparidade"], "Débito"),
    ("71", "Vendas", ["venda", "loja online", "loja fisica", "marketplace", "produto vendido", "ovoko venda", "recambio venda"], "Crédito"),
    ("72", "Prestações de serviços", ["servico", "mao de obra", "reparacao", "diagnostico"], "Crédito"),
    ("78", "Outros rendimentos e ganhos", ["ganho", "indemnizacao", "regularizacao positiva"], "Crédito"),
]

DESCRIPTION_HEADERS = ["descricao", "movimento", "documento", "produto", "artigo", "categoria", "designacao", "nome", "referencia"]
VALUE_HEADERS = ["valor", "montante", "total", "amount", "preco", "debito", "credito", "saldo"]
ENTITY_HEADERS = ["entidade", "cliente", "fornecedor", "plataforma", "canal", "marketplace", "origem"]
DATE_HEADERS = ["data", "date", "dia", "emissao", "vencimento"]
STOCK_HEADERS = ["stock", "quantidade", "qtd", "existencia"]
DEBT_HEADERS = ["valor em aberto", "montante em aberto", "saldo em aberto", "divida", "dívida", "outstandingamount"]


@app.get("/health")
def health() -> dict:
    return {"status": "ok", "service": "seo-api", "environment": APP_ENV}


async def read_validated_upload(file: UploadFile, allowed_suffixes: set[str], default_name: str) -> tuple[bytes, str, str]:
    raw_name = (file.filename or default_name).strip()
    filename = re.split(r"[/\\]", raw_name)[-1] or default_name
    name, separator, suffix = filename.lower().rpartition(".")
    if not separator or not name or suffix not in allowed_suffixes:
        formats = ", ".join(f".{item}" for item in sorted(allowed_suffixes))
        raise HTTPException(status_code=400, detail=f"Formato inválido. Formatos aceites: {formats}.")

    try:
        content = await file.read(MAX_UPLOAD_BYTES + 1)
    finally:
        await file.close()
    if not content:
        raise HTTPException(status_code=400, detail="O ficheiro está vazio.")
    if len(content) > MAX_UPLOAD_BYTES:
        limit_mb = MAX_UPLOAD_BYTES // (1024 * 1024)
        raise HTTPException(status_code=413, detail=f"Ficheiro demasiado grande. O limite é {limit_mb} MB.")

    signatures = {
        "xlsx": (b"PK\x03\x04",),
        "pdf": (b"%PDF",),
        "png": (b"\x89PNG\r\n\x1a\n",),
        "jpg": (b"\xff\xd8\xff",),
        "jpeg": (b"\xff\xd8\xff",),
    }
    expected = signatures.get(suffix)
    if expected and not any(content.startswith(signature) for signature in expected):
        raise HTTPException(status_code=400, detail="O conteúdo do ficheiro não corresponde à extensão indicada.")
    return content, filename, suffix


@app.get("/strategy/operating-model")
def strategy_model(user: dict = Depends(require_permission("dashboard:read"))) -> dict:
    return strategy_operating_model()


@app.post("/strategy/score")
def strategy_score(payload: StrategyFitRequest, user: dict = Depends(require_permission("dashboard:read"))) -> dict:
    return score_strategy_fit(payload.signals)


@app.post("/auth/register", response_model=AccountProfile)
def register(payload: RegisterRequest, request: FastAPIRequest) -> AccountProfile:
    client = request.client.host if request.client else "unknown"
    enforce_rate_limit(f"register:{client}", limit=5, window_seconds=3600)
    try:
        account = register_client(payload.name, str(payload.email), payload.password, payload.company_name)
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    record_audit_event(AuditEvent(actor=str(payload.email), action="CLIENT_REGISTERED", details="Conta cliente criada."))
    return account


@app.post("/auth/login", response_model=ChallengeResponse)
def login(payload: LoginRequest, request: FastAPIRequest) -> ChallengeResponse:
    client = request.client.host if request.client else "unknown"
    user = authenticate_user(str(payload.email), payload.password)
    if not user:
        enforce_rate_limit(f"login-failed:{client}:{str(payload.email).lower()}", limit=8, window_seconds=300)
        record_audit_event(AuditEvent(actor=str(payload.email), action="LOGIN_FAILED", details="Credenciais recusadas."))
        raise HTTPException(status_code=401, detail="Credenciais inválidas.")

    challenge_id, development_code, delivery_hint = create_mfa_challenge(user["email"])
    record_audit_event(AuditEvent(actor=user["email"], action="MFA_CHALLENGE_CREATED", details="Código temporário criado."))
    return ChallengeResponse(
        challenge_id=challenge_id,
        expires_in_seconds=MFA_TTL_SECONDS,
        delivery_hint=delivery_hint,
        development_code=development_code,
    )


@app.post("/auth/mfa", response_model=TokenResponse)
def mfa(payload: MfaRequest, request: FastAPIRequest) -> TokenResponse:
    client = request.client.host if request.client else "unknown"
    enforce_rate_limit(f"mfa:{client}:{payload.challenge_id}", limit=6, window_seconds=MFA_TTL_SECONDS)
    user = verify_mfa_challenge(payload.challenge_id, payload.code)
    if not user:
        record_audit_event(AuditEvent(actor="unknown", action="MFA_FAILED", details="Código inválido ou expirado."))
        raise HTTPException(status_code=401, detail="Código inválido ou expirado.")

    record_audit_event(AuditEvent(actor=user["email"], action="MFA_SUCCESS", details="Segunda autenticação validada."))
    return TokenResponse(access_token=create_access_token(user["email"]), account=account_profile(user))


@app.get("/me", response_model=AccountProfile)
def me(user: dict = Depends(get_current_user)) -> AccountProfile:
    return account_profile(user)


@app.post("/files/analyze")
async def analyze_file(
    file: UploadFile = File(...),
    user: dict = Depends(require_permission("files:upload")),
) -> dict:
    content, filename, suffix = await read_validated_upload(
        file,
        {"xlsx", "pdf", "csv", "txt", "xml", "jpg", "jpeg", "png"},
        "ficheiro",
    )
    rows = await run_in_threadpool(read_rows, content, suffix)
    dataset = await run_in_threadpool(rows_to_dataset, filename, rows)
    stored_file = save_uploaded_file(user["company_id"], user["email"], filename, file.content_type or "application/octet-stream", "documentos", content)
    generated_file = None
    billing_transform = None
    if suffix == "xlsx" and await run_in_threadpool(is_svp_billing_workbook, content):
        generated_content, billing_transform = await run_in_threadpool(transform_svp_billing_workbook, content, filename)
        generated_name = f"{filename.rsplit('.', 1)[0]}_separada_Coimbra_Picoto.xlsx"
        generated_file = save_uploaded_file(
            user["company_id"],
            user["email"],
            generated_name,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "faturacao-organizada",
            generated_content,
        )
    replace_operational_dataset(
        company_id=user["company_id"],
        owner_email=user["email"],
        inventory=dataset["inventory"],
        debts=dataset["debts"],
        issues=dataset["issues"],
    )
    save_operational_state(
        company_id=user["company_id"],
        owner_email=user["email"],
        source_name=filename,
        summary=dataset["summary"],
        document_intelligence=dataset["documentIntelligence"],
    )
    snapshot = save_metric_snapshot(
        company_id=user["company_id"],
        owner_email=user["email"],
        period="daily",
        label=f"Importação {filename}",
        metrics=build_company_metrics(user["company_id"]),
    )
    record_audit_event(
        AuditEvent(
            actor=user["email"],
            action="FILE_ANALYZED",
            details=f"{filename}: {dataset['summary']['rowsRead']} linhas analisadas.",
        )
    )
    row_errors = [
        {"row": index + 2, "document": document["number"], "errors": document["validations"]}
        for index, document in enumerate(dataset["documentIntelligence"]["documents"])
        if document["needsReview"]
    ][:200]
    return {
        **dataset,
        "snapshot": snapshot,
        "storedFile": stored_file,
        "generatedFile": generated_file,
        "billingTransform": billing_transform,
        "rowErrors": row_errors,
    }


@app.get("/dashboard/state")
def dashboard_state(user: dict = Depends(require_permission("dashboard:read"))) -> dict:
    state = get_operational_state(user["company_id"])
    if state:
        return state
    return {
        "sourceName": "",
        "summary": None,
        "documentIntelligence": None,
        "updatedAt": None,
    }


@app.post("/api/v1/documents/ocr")
async def ocr_document(
    company_id: str,
    file: UploadFile = File(...),
    user: dict = Depends(require_permission("documents:write")),
) -> dict:
    company = get_company(company_id)
    if not company or company_id != user["company_id"]:
        raise HTTPException(status_code=404, detail="Empresa não encontrada.")

    filename = re.split(r"[/\\]", (file.filename or "documento").strip())[-1] or "documento"
    content_type = file.content_type
    try:
        file_bytes = await file.read(MAX_OCR_FILE_SIZE + 1)
    finally:
        await file.close()
    if not file_bytes:
        raise HTTPException(status_code=422, detail="O ficheiro está vazio.")
    if len(file_bytes) > MAX_OCR_FILE_SIZE:
        raise HTTPException(status_code=413, detail="Ficheiro demasiado grande. O limite é 25 MB.")

    try:
        result = await run_in_threadpool(read_full_file, filename, content_type, file_bytes)
    except (OcrValidationError, OcrExecutionError) as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("Falha inesperada no OCR de %s para a empresa %s", filename, company_id)
        raise HTTPException(status_code=500, detail="Não foi possível ler o documento.") from exc

    record_audit_event(AuditEvent(
        actor=user["email"],
        action="document_ocr_completed",
        details=f"entity_type=document_upload; filename={filename}; pages={result['page_count']}; company_id={company_id}; uploaded_by={user['email']}",
    ))
    return {"message": "Ficheiro lido completamente.", **result}


@app.get("/reconciliation/issues")
def get_reconciliation_issues(user: dict = Depends(require_permission("reconciliation:read"))) -> list[dict]:
    return list_reconciliation_issues(user["company_id"])


@app.post("/reconciliation/issues/{issue_id}/resolve")
def resolve_issue(issue_id: int, user: dict = Depends(require_permission("reconciliation:write"))) -> dict:
    issue = resolve_reconciliation_issue(user["company_id"], issue_id)
    if not issue:
        raise HTTPException(status_code=404, detail="Pendência não encontrada.")
    record_audit_event(AuditEvent(actor=user["email"], action="ISSUE_RESOLVED", details=f"Pendência {issue_id} resolvida."))
    return issue


@app.post("/reconciliation/issues/resolve-all")
def resolve_all_issues(user: dict = Depends(require_permission("reconciliation:write"))) -> list[dict]:
    issues = resolve_all_reconciliation_issues(user["company_id"])
    record_audit_event(AuditEvent(actor=user["email"], action="ISSUES_RESOLVED", details=f"{len(issues)} pendências resolvidas."))
    return issues


@app.post("/reconciliation/import")
async def import_reconciliation_file(
    file: UploadFile = File(...),
    user: dict = Depends(require_permission("reconciliation:write")),
) -> list[dict]:
    content, filename, suffix = await read_validated_upload(file, {"csv", "txt"}, "conciliacao.csv")
    rows = await run_in_threadpool(read_rows, content, suffix)
    issues = rows_to_reconciliation_issues(rows)
    if not issues:
        raise HTTPException(status_code=422, detail="Não foram encontradas pendências de conciliação no ficheiro.")

    append_reconciliation_issues(user["company_id"], user["email"], issues)
    save_uploaded_file(user["company_id"], user["email"], filename, file.content_type or "text/plain", "conciliacao", content)
    record_audit_event(
        AuditEvent(
            actor=user["email"],
            action="RECONCILIATION_IMPORTED",
            details=f"{filename}: {len(issues)} pendências importadas.",
        )
    )
    return list_reconciliation_issues(user["company_id"])


@app.get("/inventory/items")
def get_inventory_items(user: dict = Depends(require_permission("inventory:read"))) -> list[dict]:
    return list_inventory_items(user["company_id"])


@app.post("/inventory/items/{ref}/sale")
def register_sale(ref: str, user: dict = Depends(require_permission("inventory:write"))) -> dict:
    item = register_inventory_sale(user["company_id"], ref)
    if not item:
        raise HTTPException(status_code=404, detail="Produto não encontrado.")
    record_audit_event(AuditEvent(actor=user["email"], action="INVENTORY_MOVEMENT", details=f"Saída de stock registada para {ref}."))
    return item


@app.get("/finance/debts")
def get_debts(user: dict = Depends(require_permission("finance:read"))) -> list[dict]:
    return list_debt_items(user["company_id"])


@app.post("/finance/debts/{debt_id}/pay")
def pay_debt(debt_id: int, user: dict = Depends(require_permission("finance:write"))) -> dict:
    debt = mark_debt_paid(user["company_id"], debt_id)
    if not debt:
        raise HTTPException(status_code=404, detail="Conta corrente não encontrada.")
    record_audit_event(AuditEvent(actor=user["email"], action="DEBT_MARKED_PAID", details=f"Conta corrente {debt_id} marcada como paga."))
    return debt


@app.get("/reports/executive")
def executive_report(user: dict = Depends(require_permission("reports:export"))) -> dict:
    company_id = user["company_id"]
    metrics = build_company_metrics(company_id)
    priorities = build_decision_priorities(metrics)
    return {
        "companyId": company_id,
        "generatedBy": user["email"],
        "seoIndex": metrics["seo_index"],
        "riskLevel": metrics["risk_level"],
        "capitalAtRisk": metrics["capital_at_risk"],
        "openDebt": metrics["open_debt"],
        "stalledProducts": metrics["stalled_products"],
        "criticalStock": metrics["critical_stock"],
        "unresolvedIssues": metrics["unresolved_issues"],
        "averageMargin": metrics["average_margin"],
        "priorities": priorities,
        "recommendedActions": [priority["recommendation"] for priority in priorities],
    }


@app.get("/reports/executive.pdf")
def executive_report_pdf(user: dict = Depends(require_permission("reports:export"))) -> Response:
    report = executive_report(user)
    record_audit_event(AuditEvent(actor=user["email"], action="EXECUTIVE_PDF_EXPORTED", details="Relatório executivo PDF gerado."))
    return Response(
        content=build_executive_pdf(report),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=relatorio-executivo-seo.pdf"},
    )


@app.get("/integrations/marketplaces")
def marketplace_integrations(user: dict = Depends(require_permission("dashboard:read"))) -> list[dict]:
    return [
        {"id": "ovoko", "name": "Ovoko", "status": "planned", "type": "marketplace"},
        {"id": "recambio", "name": "Recambio", "status": "planned", "type": "marketplace"},
        {"id": "woocommerce", "name": "WooCommerce", "status": "planned", "type": "store"},
        {"id": "shopify", "name": "Shopify", "status": "planned", "type": "store"},
        {"id": "moloni", "name": "Moloni", "status": "planned", "type": "accounting"},
        {"id": "primavera", "name": "Primavera", "status": "planned", "type": "erp"},
    ]


@app.get("/decision-center")
def decision_center(user: dict = Depends(require_permission("dashboard:read"))) -> dict:
    metrics = build_company_metrics(user["company_id"])
    return {
        "companyId": user["company_id"],
        "seoIndex": metrics["seo_index"],
        "riskLevel": metrics["risk_level"],
        "capitalAtRisk": metrics["capital_at_risk"],
        "priorities": build_decision_priorities(metrics),
        "positioning": "O SEO não é um dashboard. É um sistema que transforma dados operacionais em decisões financeiras prioritárias.",
    }


@app.post("/ai/analyze")
def ai_analyze(payload: AiQuestionRequest, user: dict = Depends(require_permission("dashboard:read"))) -> dict:
    metrics = build_company_metrics(user["company_id"])
    question = payload.question.strip()
    normalized = normalize(question)
    priorities = build_decision_priorities(metrics)
    intent = detect_ai_intent(normalized)
    answer, risk, confidence = compose_ai_answer(intent, metrics, priorities)
    actions = build_ai_actions(intent, metrics, priorities)
    history = get_ai_messages(user["company_id"], payload.conversation_id)
    contextual_answer = generate_contextual_answer(question, metrics, history, payload.analysis_level)
    if contextual_answer:
        answer = contextual_answer
        confidence = max(confidence, 92)
    elif payload.analysis_level == "Rápido":
        answer = answer.split(". ")[0].rstrip(".") + "."
        actions = actions[:2]
    elif payload.analysis_level == "Auditoria":
        answer += (
            f" Para auditoria: foram considerados {metrics['unresolved_issues']} registos pendentes, "
            f"{metrics['active_debts']} contas abertas e {metrics['critical_stock']} referências em stock crítico. "
            "Valide documentos de origem e exceções antes de executar alterações definitivas."
        )
        confidence = max(60, confidence - 3)
    conversation_id = save_ai_exchange(user["company_id"], user["email"], payload.conversation_id, question, answer)

    record_audit_event(AuditEvent(actor=user["email"], action="AI_ANALYSIS", details=f"Pergunta analisada: {question[:120]}"))
    return {
        "answer": answer,
        "conversationId": conversation_id,
        "provider": "openai" if contextual_answer else "analytical-fallback",
        "confidence": confidence,
        "risk": risk,
        "priorities": [priority["title"] for priority in priorities],
        "actions": actions,
        "intent": intent,
        "analysisLevel": payload.analysis_level,
        "nextQuestions": build_next_questions(intent),
        "explainability": {
            "dataSources": ["inventário persistido", "contas correntes", "conciliação", "Índice SEO Core"],
            "financialImpact": round(metrics["capital_at_risk"], 2),
            "signals": build_ai_signals(metrics),
            "method": "Regras analíticas ponderadas por impacto financeiro, urgência operacional e qualidade dos dados.",
            "humanReview": "Obrigatória para SNC, fiscalidade e decisões oficiais.",
        },
    }


@app.post("/reports/snapshots")
def create_snapshot(payload: SnapshotCreateRequest, user: dict = Depends(require_permission("reports:export"))) -> dict:
    snapshot = save_metric_snapshot(
        company_id=user["company_id"],
        owner_email=user["email"],
        period=payload.period,
        label=payload.label or f"Snapshot {payload.period}",
        metrics=build_company_metrics(user["company_id"]),
        report_date=payload.report_date.isoformat() if payload.report_date else None,
    )
    record_audit_event(AuditEvent(actor=user["email"], action="SNAPSHOT_CREATED", details=f"Snapshot {payload.period} criado."))
    return snapshot


@app.get("/reports/snapshots")
def get_snapshots(
    period: str | None = None,
    report_date: str | None = None,
    limit: int = 24,
    user: dict = Depends(require_permission("reports:export")),
) -> list[dict]:
    return list_metric_snapshots(user["company_id"], period=period, limit=limit, report_date=report_date)


@app.get("/reports/compare")
def compare_snapshots(
    period: str = "monthly",
    user: dict = Depends(require_permission("reports:export")),
) -> dict:
    snapshots = list_metric_snapshots(user["company_id"], period=period, limit=2)
    current_metrics = build_company_metrics(user["company_id"])
    if len(snapshots) < 2:
        return {
            "period": period,
            "current": current_metrics,
            "previous": None,
            "delta": {},
            "message": "Ainda não há snapshots suficientes para comparação.",
        }
    current = snapshots[0]["metrics"]
    previous = snapshots[1]["metrics"]
    return {
        "period": period,
        "current": current,
        "previous": previous,
        "delta": build_metric_delta(current, previous),
    }


@app.get("/billing/subscription")
def billing_subscription(user: dict = Depends(require_permission("billing:manage"))) -> dict:
    subscription = get_subscription(user["company_id"])
    return subscription or {"companyId": user["company_id"], "plan": "starter", "status": "trial"}


@app.post("/billing/checkout")
def billing_checkout(payload: BillingCheckoutRequest, user: dict = Depends(require_permission("billing:manage"))) -> dict:
    price_id = STRIPE_PRICES.get(payload.plan, "")
    if not STRIPE_SECRET_KEY or not price_id:
        raise HTTPException(status_code=503, detail="Stripe não configurado. Defina STRIPE_SECRET_KEY e o preço do plano.")

    checkout = create_stripe_checkout_session(
        plan=payload.plan,
        price_id=price_id,
        customer_email=user["email"],
        company_id=user["company_id"],
    )
    upsert_subscription(user["company_id"], payload.plan, "checkout_pending")
    record_audit_event(AuditEvent(actor=user["email"], action="BILLING_CHECKOUT_CREATED", details=f"Checkout criado para {payload.plan}."))
    return checkout


@app.post("/billing/webhook")
async def stripe_webhook(request: FastAPIRequest) -> dict:
    body = await request.body()
    signature = request.headers.get("stripe-signature", "")
    if not STRIPE_WEBHOOK_SECRET:
        raise HTTPException(status_code=503, detail="Webhook Stripe indisponível: segredo não configurado.")
    if not verify_stripe_signature(body, signature):
        raise HTTPException(status_code=400, detail="Assinatura Stripe inválida.")

    try:
        event = json.loads(body.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=400, detail="Evento Stripe inválido.") from exc
    event_type = event.get("type", "unknown")
    obj = event.get("data", {}).get("object", {})
    company_id = obj.get("metadata", {}).get("company_id") or "unknown"
    plan = obj.get("metadata", {}).get("plan") or "professional"
    add_payment_event(company_id, "stripe", event_type, event)

    if event_type in {"checkout.session.completed", "customer.subscription.updated", "customer.subscription.created"}:
        upsert_subscription(
            company_id=company_id,
            plan=plan,
            status=obj.get("payment_status") or obj.get("status") or "active",
            stripe_customer_id=obj.get("customer"),
            stripe_subscription_id=obj.get("subscription") or obj.get("id"),
        )
    elif event_type in {"customer.subscription.deleted", "invoice.payment_failed"}:
        upsert_subscription(company_id=company_id, plan=plan, status="past_due" if event_type.endswith("failed") else "canceled")

    return {"received": True}


@app.post("/snc/classify")
async def classify_snc_file(
    file: UploadFile = File(...),
    user: dict = Depends(require_permission("snc:read")),
) -> list[dict]:
    content, filename, suffix = await read_validated_upload(file, {"xlsx", "csv", "txt", "xml"}, "movimentos.csv")
    rows = await run_in_threadpool(read_rows, content, suffix)
    dataset = await run_in_threadpool(rows_to_dataset, filename, rows)
    save_uploaded_file(user["company_id"], user["email"], filename, file.content_type or "application/octet-stream", "snc", content)
    record_audit_event(
        AuditEvent(
            actor=user["email"],
            action="SNC_CLASSIFIED",
            details=f"{filename}: {len(dataset['classifiedMovements'])} movimentos classificados.",
        )
    )
    return dataset["classifiedMovements"]


@app.get("/audit")
def audit(limit: int = 100, user: dict = Depends(require_permission("audit:read"))) -> list[dict]:
    record_audit_event(AuditEvent(actor=user["email"], action="AUDIT_VIEWED", details=f"Consulta de {limit} eventos."))
    return list_audit_events(limit, company_id=user["company_id"])


@app.post("/audit/events", status_code=201)
def create_audit_event(payload: AuditEvent, user: dict = Depends(get_current_user)) -> dict:
    return record_audit_event(
        AuditEvent(actor=user["email"], action=payload.action[:80], details=payload.details[:500]),
        company_id=user["company_id"],
    )


@app.get("/cloud/files")
def cloud_files(limit: int = 100, user: dict = Depends(require_permission("files:upload"))) -> list[dict]:
    return list_uploaded_files(user["company_id"], limit)


@app.get("/cloud/files/{file_id}/download")
def download_cloud_file(file_id: str, user: dict = Depends(require_permission("files:upload"))) -> Response:
    stored = get_uploaded_file(user["company_id"], file_id)
    if not stored:
        raise HTTPException(status_code=404, detail="Ficheiro não encontrado.")
    safe_name = re.sub(r"[^A-Za-z0-9._-]", "_", stored["filename"]) or "documento"
    return Response(
        content=stored["content"],
        media_type=stored["content_type"],
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"', "X-Content-SHA256": stored["sha256"]},
    )


def read_rows(content: bytes, suffix: str) -> list[dict]:
    try:
        if suffix in {"jpg", "jpeg", "png"}:
            text = extract_image_text(content)
            if not text.strip():
                raise ValueError("OCR local não reconheceu texto na imagem")
            return pdf_text_to_rows(text)

        if suffix == "pdf":
            text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(content)).pages)
            if not text.strip():
                text = extract_scanned_pdf_text(content)
            if not text.strip():
                raise ValueError("OCR local não reconheceu texto no PDF")
            return pdf_text_to_rows(text)

        if suffix == "xml":
            root = ET.fromstring(content)
            rows = []
            for element in root.iter():
                local_tag = element.tag.split("}")[-1].lower()
                if local_tag not in {"invoice", "creditnote", "debitnote", "document"}:
                    continue
                row = {}
                for child in element.iter():
                    if list(child) or not child.text or not child.text.strip():
                        continue
                    key = child.tag.split("}")[-1]
                    row.setdefault(key, child.text.strip())
                normalized_keys = {normalize(key) for key in row}
                if {"invoiceno", "documentnumber", "id"} & normalized_keys:
                    rows.append(row)
            if not rows and root.tag.split("}")[-1].lower() in {"invoice", "creditnote", "debitnote"}:
                row = {}
                for child in root.iter():
                    if not list(child) and child.text and child.text.strip():
                        row.setdefault(child.tag.split("}")[-1], child.text.strip())
                if row:
                    rows.append(row)
            if not rows:
                raise ValueError("XML sem documentos reconhecíveis")
            return rows

        if suffix == "xlsx":
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            rows: list[dict] = []
            for worksheet in workbook.worksheets:
                values = worksheet.iter_rows(values_only=True)
                rows.extend(normalize_excel_sheet(values))
            workbook.close()
            return rows

        text = content.decode("utf-8-sig")
        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,	,")
        except csv.Error:
            dialect = csv.excel
        return list(csv.DictReader(text.splitlines(), dialect=dialect))
    except Exception as exc:
        detail = "Não foi possível ler o ficheiro. Confirme o formato e os cabeçalhos."
        if suffix in {"pdf", "jpg", "jpeg", "png"} and "OCR local" in str(exc):
            detail = "O OCR local não conseguiu reconhecer os dados do documento. Use uma imagem mais nítida ou envie para revisão."
        raise HTTPException(status_code=400, detail=detail) from exc


def pdf_text_to_rows(text: str) -> list[dict]:
    compact = re.sub(r"[ \t]+", " ", text)
    number = first_pdf_match(compact, [r"(?:fatura|factura|invoice|documento)\s*(?:n[.ºo°]*|no)?\s*[:#-]?\s*([A-Z0-9./-]+)"])
    date = first_pdf_match(compact, [r"(?:data|date|emiss[aã]o)\s*[:#-]?\s*(\d{1,4}[-/]\d{1,2}[-/]\d{1,4})"])
    entity = first_pdf_match(compact, [r"(?:fornecedor|supplier|emitente)\s*[:#-]?\s*([^\n]{2,100})"])
    total = first_pdf_match(compact, [r"(?:^|\n)\s*(?:total a pagar|total|payable amount)\s*[:€ ]+\s*([\d.,]+)"])
    vat = first_pdf_match(compact, [r"(?:^|\n)\s*(?:total iva|iva|vat|tax payable)\s*[:€ ]+\s*([\d.,]+)"])
    net = first_pdf_match(compact, [r"(?:^|\n)\s*(?:valor sem iva|subtotal|net total|base tribut[aá]vel)\s*[:€ ]+\s*([\d.,]+)"])
    if not total:
        raise ValueError("PDF sem total reconhecível")
    description = first_pdf_match(compact, [r"(?:descri[cç][aã]o|description)\s*[:#-]?\s*([^\n]{2,160})"]) or "Fatura importada de PDF"
    return [{"documento": number or "PDF-1", "data": date or "-", "entidade": entity or "Não identificado", "descricao": description, "valor sem iva": net or "0", "iva": vat or "0", "total": total, "estado": "Desconhecido"}]


def normalize_excel_sheet(values) -> list[dict]:
    buffered_rows = []
    for _ in range(30):
        try:
            buffered_rows.append(next(values))
        except StopIteration:
            break
    non_empty_rows = [(index, row) for index, row in enumerate(buffered_rows) if any(not is_empty(value) for value in row)]
    if not non_empty_rows:
        return []
    known_headers = DESCRIPTION_HEADERS + VALUE_HEADERS + ENTITY_HEADERS + DATE_HEADERS + STOCK_HEADERS + DEBT_HEADERS + [
        "numero", "documento", "fatura", "iva", "valor sem iva", "taxa iva", "nif", "iban", "estado", "pago",
    ]
    candidates = non_empty_rows[:30]
    header_index, header_row = max(candidates, key=lambda item: score_excel_header(item[1], known_headers))
    if score_excel_header(header_row, known_headers) == 0:
        header_index, header_row = non_empty_rows[0]
    header = dedupe_headers(header_row)
    normalized_rows = []
    for row in chain(buffered_rows[header_index + 1 :], values):
        if not any(not is_empty(value) for value in row):
            continue
        payload = {header[index]: value for index, value in enumerate(row[: len(header)]) if not is_empty(value)}
        if payload and not is_probable_total_row(payload):
            normalized_rows.append(payload)
    return normalized_rows


def score_excel_header(row: tuple, known_headers: list[str]) -> int:
    cells = [normalize(str(value)) for value in row if isinstance(value, str) and value.strip()]
    return sum(3 if any(cell == normalize(candidate) for candidate in known_headers) else 1 if any(normalize(candidate) in cell for candidate in known_headers) else 0 for cell in cells)


def dedupe_headers(row: tuple) -> list[str]:
    result = []
    counts: dict[str, int] = {}
    for index, value in enumerate(row):
        base = str(value or "").strip() or f"coluna_{index + 1}"
        counts[base] = counts.get(base, 0) + 1
        result.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return result


def is_probable_total_row(row: dict) -> bool:
    texts = [normalize(str(value)).strip() for value in row.values() if isinstance(value, str)]
    return any(value in {"total", "total geral", "subtotal", "totais"} for value in texts) and not any("fatura" in value or "invoice" in value for value in texts)


def first_pdf_match(text: str, patterns: list[str]) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE | re.MULTILINE)
        if match:
            return match.group(1).strip()
    return ""


def rows_to_reconciliation_issues(rows: list[dict]) -> list[dict]:
    base_id = int(datetime.now(timezone.utc).timestamp() * 1000) % 100_000_000
    issues = []
    for index, row in enumerate(rows[:500]):
        normalized_row = {normalize(str(key)): value for key, value in row.items()}
        document = (
            pick_value(normalized_row, ["documento", "doc", "fatura", "referencia", "numero"])
            or longest_text_value(row)
            or f"IMP-{index + 1}"
        )
        source = pick_value(normalized_row, ENTITY_HEADERS) or pick_value(normalized_row, ["origem", "fonte"]) or "Importado"
        amount = parse_number(pick_value(normalized_row, VALUE_HEADERS) or last_numeric_value(row))
        issue = (
            pick_value(normalized_row, ["erro", "pendencia", "inconsistencia", "problema", "observacao"])
            or "Movimento importado para validação"
        )
        status = normalize(str(pick_value(normalized_row, ["estado", "status"]) or "Rever"))
        normalized_status = "Classificar" if "class" in status else "Alerta" if "alert" in status else "Rever"

        if is_empty(document) and is_empty(source) and amount == 0:
            continue

        issues.append(
            {
                "id": base_id + index,
                "document": str(document)[:120],
                "source": str(source)[:120],
                "value": f"{amount:.2f} EUR",
                "issue": str(issue)[:180],
                "status": normalized_status,
            }
        )
    return issues


def build_simple_pdf(lines: list[str]) -> bytes:
    escaped_lines = [line.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)") for line in lines]
    text_commands = ["BT", "/F1 12 Tf", "50 790 Td"]
    for index, line in enumerate(escaped_lines):
        if index > 0:
            text_commands.append("0 -18 Td")
        text_commands.append(f"({line}) Tj")
    text_commands.append("ET")
    stream = "\n".join(text_commands)
    objects = [
        "1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj",
        "2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj",
        "3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj",
        "4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj",
        f"5 0 obj << /Length {len(stream.encode('latin-1', errors='replace'))} >> stream\n{stream}\nendstream endobj",
    ]
    pdf = "%PDF-1.4\n"
    offsets = []
    for obj in objects:
        offsets.append(len(pdf.encode("latin-1", errors="replace")))
        pdf += obj + "\n"
    xref_offset = len(pdf.encode("latin-1", errors="replace"))
    pdf += f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n"
    pdf += "".join(f"{offset:010d} 00000 n \n" for offset in offsets)
    pdf += f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF"
    return pdf.encode("latin-1", errors="replace")


def build_executive_pdf(report: dict) -> bytes:
    commands: list[str] = []

    def color(hex_color: str) -> str:
        value = hex_color.lstrip("#")
        red = int(value[0:2], 16) / 255
        green = int(value[2:4], 16) / 255
        blue = int(value[4:6], 16) / 255
        return f"{red:.3f} {green:.3f} {blue:.3f}"

    def rect(x: int, y: int, width: int, height: int, fill: str) -> None:
        commands.append(f"{color(fill)} rg {x} {y} {width} {height} re f")

    def text(x: int, y: int, value: str, size: int = 11, fill: str = "#111827") -> None:
        safe = str(value).encode("latin-1", errors="replace").decode("latin-1")
        safe = safe.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        commands.append(f"BT {color(fill)} rg /F1 {size} Tf 1 0 0 1 {x} {y} Tm ({safe}) Tj ET")

    def wrapped_text(x: int, y: int, value: str, width: int = 70, size: int = 10, fill: str = "#334155") -> int:
        words = str(value).split()
        lines: list[str] = []
        current = ""
        for word in words:
            next_line = f"{current} {word}".strip()
            if len(next_line) > width and current:
                lines.append(current)
                current = word
            else:
                current = next_line
        if current:
            lines.append(current)
        for index, line in enumerate(lines[:3]):
            text(x, y - index * 13, line, size, fill)
        return y - min(len(lines), 3) * 13

    rect(0, 0, 595, 842, "#f6f7fb")
    rect(0, 710, 595, 132, "#071527")
    text(42, 792, "SEO CORE", 13, "#93c5fd")
    text(42, 758, "Relatorio Executivo", 28, "#ffffff")
    text(42, 733, "Centro de Decisao Operacional com IA explicavel", 12, "#cbd5e1")
    text(420, 792, f"Indice SEO {report['seoIndex']}/100", 16, "#ffffff")
    text(420, 765, f"Risco: {report['riskLevel']}", 12, "#fef3c7")

    kpis = [
        ("Capital em risco", f"{report['capitalAtRisk']:.2f} EUR"),
        ("Saldos em aberto", f"{report['openDebt']:.2f} EUR"),
        ("Produtos parados", str(report["stalledProducts"])),
        ("Pendencias", str(report["unresolvedIssues"])),
    ]
    for index, (label, value) in enumerate(kpis):
        x = 42 + index * 128
        rect(x, 630, 116, 58, "#ffffff")
        text(x + 12, 670, label, 8, "#64748b")
        text(x + 12, 646, value, 15, "#0f172a")

    text(42, 590, "Resumo Executivo da IA", 17, "#0f172a")
    wrapped_text(
        42,
        568,
        (
            f"A operacao apresenta indice {report['seoIndex']}/100 e risco {str(report['riskLevel']).lower()}. "
            f"O impacto financeiro prioritario soma {report['capitalAtRisk']:.2f} EUR entre saldos, stock e pendencias."
        ),
        92,
        10,
        "#334155",
    )

    text(42, 505, "Top 3 decisoes prioritarias", 16, "#0f172a")
    y = 476
    for index, priority in enumerate(report["priorities"][:3], start=1):
        rect(42, y - 45, 510, 58, "#ffffff")
        text(58, y - 2, f"{index}. {priority['title']}", 12, "#0f172a")
        text(390, y - 2, f"{priority['financialImpact']:.2f} EUR", 11, "#0f766e")
        wrapped_text(58, y - 21, priority["recommendation"], 78, 9, "#475569")
        y -= 72

    text(42, 245, "Plano de acao executivo", 16, "#0f172a")
    y = 218
    for action in report["recommendedActions"][:4]:
        text(58, y, f"- {action}", 10, "#334155")
        y -= 18

    rect(42, 72, 510, 54, "#eaf2ff")
    text(58, 104, "Nota de conformidade", 11, "#0f172a")
    wrapped_text(
        58,
        86,
        "As sugestoes da IA apoiam decisao de gestao e exigem validacao humana para SNC, fiscalidade e reporte oficial.",
        88,
        9,
        "#334155",
    )

    stream = "\n".join(commands)
    objects = [
        "1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj",
        "2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj",
        "3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj",
        "4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj",
        f"5 0 obj << /Length {len(stream.encode('latin-1', errors='replace'))} >> stream\n{stream}\nendstream endobj",
    ]
    pdf = "%PDF-1.4\n"
    offsets = []
    for obj in objects:
        offsets.append(len(pdf.encode("latin-1", errors="replace")))
        pdf += obj + "\n"
    xref_offset = len(pdf.encode("latin-1", errors="replace"))
    pdf += f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n"
    pdf += "".join(f"{offset:010d} 00000 n \n" for offset in offsets)
    pdf += f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF"
    return pdf.encode("latin-1", errors="replace")


def build_company_metrics(company_id: str) -> dict:
    inventory = list_inventory_items(company_id)
    debts = list_debt_items(company_id)
    issues = list_reconciliation_issues(company_id)

    unresolved_issues = len([item for item in issues if item["status"] != "Resolvido"])
    active_debts = [item for item in debts if item["state"] != "Pago"]
    open_debt = sum(item["amount"] for item in active_debts)
    stalled_products = len([item for item in inventory if item["lastSaleDays"] > 90])
    critical_stock = len([item for item in inventory if item["stock"] <= 1])
    average_margin = round(sum(item["margin"] for item in inventory) / len(inventory), 1) if inventory else 0
    capital_at_risk = open_debt + stalled_products * 75 + unresolved_issues * 320
    unresolved_rate = unresolved_issues / len(issues) if issues else 0
    overdue_count = len([item for item in active_debts if item["state"] == "Em atraso"])
    overdue_rate = overdue_count / len(debts) if debts else 0
    divergence_count = len([item for item in inventory if item.get("differenceQuantity", 0) != 0])
    divergence_rate = divergence_count / len(inventory) if inventory else 0
    critical_rate = critical_stock / len(inventory) if inventory else 0
    stalled_rate = stalled_products / len(inventory) if inventory else 0
    operational_penalty = (
        unresolved_rate * 35
        + overdue_rate * 25
        + divergence_rate * 20
        + critical_rate * 10
        + stalled_rate * 10
    )
    seo_index = max(0, min(100, round(100 - operational_penalty)))
    risk_score = unresolved_issues * 3 + len(active_debts) * 2 + stalled_products + critical_stock
    risk_level = "Elevado" if risk_score >= 10 else "Médio" if risk_score >= 4 else "Baixo"

    return {
        "seo_index": seo_index,
        "risk_level": risk_level,
        "capital_at_risk": round(capital_at_risk, 2),
        "unresolved_issues": unresolved_issues,
        "active_debts": len(active_debts),
        "open_debt": round(open_debt, 2),
        "stalled_products": stalled_products,
        "critical_stock": critical_stock,
        "average_margin": average_margin,
    }


def build_metric_delta(current: dict, previous: dict) -> dict:
    numeric_keys = [
        "seo_index",
        "capital_at_risk",
        "unresolved_issues",
        "active_debts",
        "open_debt",
        "stalled_products",
        "critical_stock",
        "average_margin",
    ]
    delta = {}
    for key in numeric_keys:
        current_value = current.get(key, 0)
        previous_value = previous.get(key, 0)
        if isinstance(current_value, (int, float)) and isinstance(previous_value, (int, float)):
            delta[key] = round(current_value - previous_value, 2)
    return delta


def create_stripe_checkout_session(plan: str, price_id: str, customer_email: str, company_id: str) -> dict:
    payload = {
        "mode": "subscription",
        "line_items[0][price]": price_id,
        "line_items[0][quantity]": "1",
        "customer_email": customer_email,
        "success_url": f"{FRONTEND_URL}?billing=success",
        "cancel_url": f"{FRONTEND_URL}?billing=cancelled",
        "metadata[company_id]": company_id,
        "metadata[plan]": plan,
        "subscription_data[metadata][company_id]": company_id,
        "subscription_data[metadata][plan]": plan,
    }
    request = Request(
        "https://api.stripe.com/v1/checkout/sessions",
        data=urlencode(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {STRIPE_SECRET_KEY}",
            "Content-Type": "application/x-www-form-urlencoded",
        },
        method="POST",
    )
    try:
        with urlopen(request, timeout=15) as response:
            data = json.loads(response.read().decode("utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=502, detail="Não foi possível criar checkout no Stripe.") from exc
    return {"checkoutUrl": data["url"], "sessionId": data["id"]}


def verify_stripe_signature(body: bytes, signature_header: str) -> bool:
    values = {}
    for part in signature_header.split(","):
        if "=" in part:
            key, value = part.split("=", 1)
            values.setdefault(key, []).append(value)
    timestamp = values.get("t", [""])[0]
    signatures = values.get("v1", [])
    if not timestamp or not signatures:
        return False
    try:
        if abs(int(datetime.now(timezone.utc).timestamp()) - int(timestamp)) > 300:
            return False
    except ValueError:
        return False
    signed_payload = f"{timestamp}.{body.decode('utf-8')}".encode("utf-8")
    expected = hmac.new(STRIPE_WEBHOOK_SECRET.encode("utf-8"), signed_payload, hashlib.sha256).hexdigest()
    return any(hmac.compare_digest(expected, signature) for signature in signatures)


def build_decision_priorities(metrics: dict) -> list[dict]:
    priorities = [
        {
            "title": "Fecho mensal com pendências",
            "severity": "Crítica" if metrics["unresolved_issues"] else "Monitorização",
            "financialImpact": round(metrics["unresolved_issues"] * 320, 2),
            "recommendation": f"Resolver {metrics['unresolved_issues']} pendências antes do fecho mensal.",
            "target": "conciliacao",
        },
        {
            "title": "Capital parado em inventário",
            "severity": "Atenção" if metrics["stalled_products"] else "Monitorização",
            "financialImpact": round(metrics["stalled_products"] * 75, 2),
            "recommendation": f"Rever preço, exposição e rotação de {metrics['stalled_products']} produtos parados.",
            "target": "inventario",
        },
        {
            "title": "Saldos por regularizar",
            "severity": "Atenção" if metrics["open_debt"] else "Monitorização",
            "financialImpact": metrics["open_debt"],
            "recommendation": f"Priorizar cobrança e confirmação de {metrics['active_debts']} contas em aberto.",
            "target": "financeiro",
        },
    ]
    return sorted(priorities, key=lambda item: item["financialImpact"], reverse=True)[:3]


def detect_ai_intent(normalized_question: str) -> str:
    intent_keywords = {
        "inventario": ["stock", "inventario", "parado", "rotacao", "ruptura", "sku", "produto"],
        "financeiro": ["lucro", "resultado", "margem", "ebitda", "venda", "receita", "despesa", "custo"],
        "cobranca": ["cliente", "divida", "cobranca", "atraso", "receber", "saldo"],
        "snc": ["snc", "conta", "classific", "contabil", "fiscal", "iva"],
        "prioridades": ["prioridade", "priorizar", "acao", "decidir", "risco", "urgente"],
    }
    scored = [
        (sum(1 for keyword in keywords if keyword in normalized_question), intent)
        for intent, keywords in intent_keywords.items()
    ]
    score, intent = max(scored, key=lambda item: item[0])
    return intent if score > 0 else "executivo"


def compose_ai_answer(intent: str, metrics: dict, priorities: list[dict]) -> tuple[str, str, int]:
    top_priority = priorities[0] if priorities else {"title": "qualidade dos dados", "financialImpact": 0}
    if intent == "inventario":
        risk = "Médio" if metrics["stalled_products"] or metrics["critical_stock"] else "Baixo"
        answer = (
            f"O inventário tem {metrics['stalled_products']} produtos parados e {metrics['critical_stock']} referências em stock crítico. "
            f"O capital operacional em risco é estimado em {metrics['capital_at_risk']:.2f} EUR. "
            "A ação com maior retorno é rever primeiro os SKUs sem rotação e separar produtos para desconto, destaque em marketplace ou reposição."
        )
        return answer, risk, confidence_from_metrics(metrics, 5)
    if intent == "cobranca":
        risk = "Elevado" if metrics["open_debt"] > 0 else "Baixo"
        answer = (
            f"As contas em aberto somam {metrics['open_debt']:.2f} EUR em {metrics['active_debts']} saldos ativos. "
            "A recomendação é ordenar por antiguidade, tratar primeiro saldos vencidos e marcar pagamentos apenas após confirmação bancária ou documental."
        )
        return answer, risk, confidence_from_metrics(metrics, 7)
    if intent == "snc":
        risk = "Médio" if metrics["unresolved_issues"] else "Baixo"
        answer = (
            f"Há {metrics['unresolved_issues']} pendências de classificação ou conciliação. "
            "A IA pode sugerir contas SNC por palavras-chave e contexto, mas deve mostrar a razão da classificação e manter validação humana para efeitos contabilísticos e fiscais."
        )
        return answer, risk, confidence_from_metrics(metrics, -2)
    if intent == "financeiro":
        risk = metrics["risk_level"]
        answer = (
            f"O Índice SEO Core está em {metrics['seo_index']}/100, com margem média operacional de {metrics['average_margin']}%. "
            f"O maior impacto financeiro imediato está em {top_priority['title'].lower()}, estimado em {top_priority['financialImpact']:.2f} EUR."
        )
        return answer, risk, confidence_from_metrics(metrics, 4)
    if intent == "prioridades":
        risk = metrics["risk_level"]
        answer = (
            f"A decisão prioritária é {top_priority['title'].lower()}. "
            f"Ela concentra {top_priority['financialImpact']:.2f} EUR de impacto estimado e deve ser tratada antes das restantes ações operacionais."
        )
        return answer, risk, confidence_from_metrics(metrics, 6)

    answer = (
        f"A operação está com Índice SEO Core de {metrics['seo_index']}/100 e risco {metrics['risk_level'].lower()}. "
        f"Foram detetadas {metrics['unresolved_issues']} pendências, {metrics['active_debts']} contas em aberto e {metrics['stalled_products']} produtos parados. "
        f"A primeira recomendação é: {top_priority['recommendation']}"
    )
    return answer, metrics["risk_level"], confidence_from_metrics(metrics, 0)


def confidence_from_metrics(metrics: dict, adjustment: int = 0) -> int:
    data_points = metrics["unresolved_issues"] + metrics["active_debts"] + metrics["stalled_products"] + metrics["critical_stock"]
    base = 76 if data_points == 0 else 88
    return max(60, min(96, base + adjustment))


def build_ai_actions(intent: str, metrics: dict, priorities: list[dict]) -> list[str]:
    base_actions = [priority["recommendation"] for priority in priorities[:3]]
    intent_actions = {
        "inventario": [
            "Ordenar inventário por dias sem venda e margem.",
            "Separar produtos para desconto, destaque ou reposição.",
        ],
        "cobranca": [
            "Criar lista de cobrança por antiguidade.",
            "Confirmar pagamentos antes de marcar saldos como pagos.",
        ],
        "snc": [
            "Aprovar automaticamente apenas classificações com confiança alta.",
            "Rever manualmente movimentos fiscais, IVA e contas genéricas.",
        ],
        "financeiro": [
            "Comparar margem real por canal antes de decidir preço.",
            "Exportar relatório executivo após resolver pendências.",
        ],
        "prioridades": [
            "Resolver a prioridade de maior impacto antes das restantes.",
            "Registar a decisão no histórico de auditoria.",
        ],
    }
    actions = intent_actions.get(intent, []) + base_actions
    if metrics["seo_index"] < 70:
        actions.insert(0, "Tratar pendências críticas antes de exportar relatório.")
    return dedupe_text(actions)[:5]


def build_ai_signals(metrics: dict) -> list[str]:
    return [
        f"Índice SEO Core: {metrics['seo_index']}/100",
        f"Risco operacional: {metrics['risk_level']}",
        f"Capital em risco: {metrics['capital_at_risk']:.2f} EUR",
        f"Pendências abertas: {metrics['unresolved_issues']}",
        f"Contas em aberto: {metrics['active_debts']}",
        f"Produtos parados: {metrics['stalled_products']}",
    ]


def build_next_questions(intent: str) -> list[str]:
    suggestions = {
        "inventario": [
            "Quais produtos devo liquidar primeiro?",
            "Onde há risco de ruptura de stock?",
            "Que SKUs têm pior margem?",
        ],
        "cobranca": [
            "Quais saldos devo cobrar primeiro?",
            "Qual é o valor em atraso por antiguidade?",
            "Que clientes aumentam o risco financeiro?",
        ],
        "snc": [
            "Por que a IA sugeriu esta conta SNC?",
            "Quais movimentos têm baixa confiança?",
            "Que classificações exigem validação humana?",
        ],
        "financeiro": [
            "Onde estou a perder margem?",
            "Qual canal gera maior rentabilidade?",
            "Que despesas mais afetam o resultado?",
        ],
    }
    return suggestions.get(intent, [
        "O que devo resolver primeiro?",
        "Qual é o maior impacto financeiro?",
        "Que risco impede o fecho mensal?",
    ])


def dedupe_text(items: list[str]) -> list[str]:
    seen = set()
    result = []
    for item in items:
        key = normalize(item)
        if key not in seen:
            seen.add(key)
            result.append(item)
    return result


def rows_to_dataset(source_name: str, rows: list[dict]) -> dict:
    movements = []
    inventory = []
    debts = []
    normalized_rows = []

    for index, row in enumerate(rows):
        normalized_row = {normalize(str(key)): value for key, value in row.items()}
        normalized_rows.append(normalized_row)
        movement = row_to_movement(row, normalized_row, index)
        if movement:
            movements.append(movement)

        stock_item = row_to_inventory(row, normalized_row, index)
        if stock_item:
            inventory.append(stock_item)

        debt = row_to_debt(row, normalized_row, index)
        if debt:
            debts.append(debt)

    if not movements:
        raise HTTPException(status_code=422, detail="A IA não encontrou movimentos com descrição e valor no ficheiro.")

    summary = build_summary(source_name, len(rows), movements)
    document_intelligence = build_document_intelligence(source_name, rows, normalized_rows)
    return {
        "summary": summary,
        "classifiedMovements": movements[:1000],
        "inventory": dedupe_by(inventory, "ref")[:250],
        "debts": dedupe_debts(debts)[:250],
        "issues": build_issues(movements, document_intelligence, inventory, debts),
        "documentIntelligence": document_intelligence,
    }


def build_document_intelligence(source_name: str, rows: list[dict], normalized_rows: list[dict] | None = None) -> dict:
    documents = []
    seen = set()
    duplicate_count = 0
    today = datetime.now(timezone.utc).date()
    for index, row in enumerate(rows):
        normalized = normalized_rows[index] if normalized_rows is not None else {normalize(str(key)): value for key, value in row.items()}
        number = pick_value(normalized, ["numero documento", "documento", "fatura", "invoice no", "invoiceno", "documentnumber", "referencia"]) or f"DOC-{index + 1}"
        entity = pick_value(normalized, ENTITY_HEADERS) or pick_value(normalized, ["customerid", "supplierid", "nome"]) or "Não identificado"
        date_value = pick_value(normalized, DATE_HEADERS + ["invoicedate", "documentdate", "systementrydate"])
        date = normalize_date(date_value)
        total_value = pick_value(normalized, ["grossamount", "grosstotal", "payableamount", "documenttotal", "valor total", "total"])
        vat = parse_number(pick_value(normalized, ["iva", "taxpayable", "taxamount", "valor iva"]))
        net = parse_number(pick_value(normalized, ["valor sem iva", "liquido", "nettotal", "netamount"]))
        total = parse_number(total_value)
        if is_empty(total_value) and net == 0 and vat == 0:
            total = parse_number(pick_value(normalized, VALUE_HEADERS) or last_numeric_value(row))
        vat_rate = abs(parse_number(pick_value(normalized, ["taxa iva", "tax rate", "taxpercentage", "percentagem iva"])))
        calculations = []
        if total == 0 and (net != 0 or vat != 0):
            total = net + vat
            calculations.append("Total recalculado a partir do valor líquido e IVA")
        elif vat == 0 and net != 0 and total != 0:
            vat = total - net
            calculations.append("IVA recalculado a partir do valor líquido e total")
        elif net == 0 and total != 0 and vat != 0:
            net = total - vat
            calculations.append("Valor líquido recalculado a partir do total e IVA")
        elif net == 0 and vat == 0 and total != 0 and vat_rate > 0:
            net = total / (1 + vat_rate / 100)
            vat = total - net
            calculations.append(f"Líquido e IVA recalculados com taxa de {vat_rate:g}%")
        text = normalize(" ".join(str(value) for value in row.values()))
        doc_type = "Nota de crédito" if "nota de credito" in text or "creditnote" in text or str(number).upper().startswith("NC") else "Nota de débito" if "nota de debito" in text else "Fatura-recibo" if "fatura recibo" in text else "Recibo" if "recibo" in text else "Fatura"
        if doc_type == "Nota de crédito":
            net, vat, total = -abs(net), -abs(vat), -abs(total)
        validations = list(calculations)
        if abs((net + vat) - total) > 0.02:
            validations.append("Total não corresponde à soma do valor líquido e IVA")
        key = f"{normalize(str(number))}|{round(total, 2)}|{normalize(str(entity))}"
        if key in seen:
            validations.append("Possível documento duplicado")
            duplicate_count += 1
        seen.add(key)
        parsed_date = None
        for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                parsed_date = datetime.strptime(str(date)[:10], pattern).date()
                break
            except ValueError:
                pass
        if date not in {"", "-"} and parsed_date is None:
            validations.append("Data inválida ou não reconhecida")
        state = "Pago" if any(word in text for word in ["pago", "liquidado", "paid"]) else "Vencido" if parsed_date and parsed_date < today and any(word in text for word in ["pendente", "vencido", "por pagar", "open"]) else "Pendente" if any(word in text for word in ["pendente", "por pagar", "open"]) else "Desconhecido"
        blocking_validations = [item for item in validations if "não corresponde" in item.lower() or "inválid" in item.lower() or "duplicado" in item.lower()]
        confidence = max(45, 98 - len(blocking_validations) * 18 - (10 if entity == "Não identificado" else 0))
        documents.append({"id": f"doc-{index + 1}", "number": str(number)[:80], "date": date, "entity": str(entity)[:120], "documentType": doc_type, "financialState": state, "netAmount": round(net, 2), "vatAmount": round(vat, 2), "totalAmount": round(total, 2), "confidence": confidence, "validations": validations, "needsReview": bool(blocking_validations) or confidence < 75})
    total_net = round(sum(item["netAmount"] for item in documents), 2)
    total_vat = round(sum(item["vatAmount"] for item in documents), 2)
    total_amount = round(sum(item["totalAmount"] for item in documents), 2)
    review = sum(1 for item in documents if item["needsReview"])
    corrected = sum(1 for item in documents if any("recalculado" in validation.lower() for validation in item["validations"]))
    return {"sourceFormat": source_name.rsplit(".", 1)[-1].upper(), "documents": documents[:1000], "totals": {"net": total_net, "vat": total_vat, "total": total_amount}, "stats": {"processed": len(documents), "valid": len(documents) - review, "review": review, "duplicates": duplicate_count, "overdue": sum(1 for item in documents if item["financialState"] == "Vencido"), "corrected": corrected}, "auditTrail": [f"{len(documents)} documentos extraídos de {source_name}", f"{corrected} documentos corrigidos ou recalculados automaticamente", "Classificação, cálculo fiscal e deteção de duplicados executados", "Resultados mantidos para validação humana antes de uso contabilístico"]}


def row_to_movement(row: dict, normalized_row: dict, index: int) -> dict | None:
    description = pick_value(normalized_row, DESCRIPTION_HEADERS) or longest_text_value(row)
    total_value = pick_value(normalized_row, ["grossamount", "grosstotal", "payableamount", "documenttotal", "valor total", "total"])
    net = parse_number(pick_value(normalized_row, ["valor sem iva", "liquido", "nettotal", "netamount"]))
    vat = parse_number(pick_value(normalized_row, ["iva", "taxpayable", "taxamount", "valor iva"]))
    amount = parse_number(total_value)
    if is_empty(total_value):
        amount = net + vat if net != 0 or vat != 0 else parse_number(pick_value(normalized_row, VALUE_HEADERS) or last_numeric_value(row))
    if not description or amount == 0:
        return None

    entity = pick_value(normalized_row, ENTITY_HEADERS) or "Não identificado"
    date = normalize_date(pick_value(normalized_row, DATE_HEADERS))
    account_code, account_name, movement_type, confidence, reason = classify_movement(f"{description} {entity}", amount)
    return {
        "id": index + 1,
        "date": date,
        "description": str(description)[:180],
        "entity": str(entity)[:120],
        "amount": amount,
        "accountCode": account_code,
        "accountName": account_name,
        "movementType": movement_type,
        "confidence": confidence,
        "reason": reason,
    }


SVP_UNIT_RULES = {
    "CUSA": ("Coimbra", "Usado"),
    "CNOV": ("Coimbra", "Novo"),
    "P NOV": ("Picoto", "Novo"),
    "PUSA": ("Picoto", "Usado"),
}


def row_to_inventory(row: dict, normalized_row: dict, index: int) -> dict | None:
    stock_value = pick_value(normalized_row, STOCK_HEADERS)
    stock = parse_number(stock_value)
    product = pick_value(normalized_row, DESCRIPTION_HEADERS) or longest_text_value(row)
    if not product or is_empty(stock_value):
        return None

    last_sale_days = int(abs(parse_number(pick_value(normalized_row, ["ultima venda", "dias", "rotacao"]))))
    margin = int(abs(parse_number(pick_value(normalized_row, ["margem", "margin"])))) or 20
    ref = pick_value(normalized_row, ["referencia", "ref", "sku", "codigo"]) or f"AUTO-{index + 1}"
    raw_code = str(pick_value(normalized_row, ["codigo unidade", "codigo interno", "serie", "unidade", "armazem"]) or "").strip().upper()
    normalized_code = re.sub(r"\s+", " ", raw_code)
    unit, stock_type = SVP_UNIT_RULES.get(normalized_code, ("Não identificado", "Não identificado"))
    system_value = pick_value(normalized_row, ["quantidade sistema", "stock sistema", "qtd sistema"])
    physical_value = pick_value(normalized_row, ["quantidade fisica", "stock fisico", "qtd fisica"])
    system_quantity = int(round(parse_number(system_value))) if not is_empty(system_value) else int(round(stock))
    physical_quantity = int(round(parse_number(physical_value))) if not is_empty(physical_value) else int(round(stock))
    difference = physical_quantity - system_quantity
    unit_cost = abs(parse_number(pick_value(normalized_row, ["custo unitario", "custo", "preco custo"])))
    confidence = 100 if normalized_code in SVP_UNIT_RULES else 70 if not normalized_code else 45
    validation_state = "Validado por regra empresarial" if normalized_code in SVP_UNIT_RULES else "Revisão necessária" if normalized_code else "Validado"
    alerts = []
    if stock < 0:
        alerts.append("Stock negativo")
    if difference:
        alerts.append("Divergência de inventário")
    if stock <= 1:
        alerts.append("Stock crítico")
    elif last_sale_days > 90:
        alerts.append("Parado")
    movement_text = normalize(" ".join(str(value) for value in row.values()))
    if any(word in movement_text for word in ["sucata", "abate", "desmantel"]):
        movement_type = "Sucata"
        movement_quantity = -abs(int(round(parse_number(pick_value(normalized_row, ["quantidade movimento", "quantidade", "qtd"]) or 1))))
    elif any(word in movement_text for word in ["venda", "saida", "saída", "vendido"]):
        movement_type = "Venda"
        movement_quantity = -abs(int(round(parse_number(pick_value(normalized_row, ["quantidade movimento", "quantidade", "qtd"]) or 1))))
    elif any(word in movement_text for word in ["compra", "entrada", "rececao", "receção"]):
        movement_type = "Compra"
        movement_quantity = abs(int(round(parse_number(pick_value(normalized_row, ["quantidade movimento", "quantidade", "qtd"]) or 1))))
    else:
        movement_type = "Existente"
        movement_quantity = physical_quantity
    return {
        "ref": str(ref)[:40],
        "product": str(product)[:140],
        "stock": int(round(stock)),
        "lastSaleDays": last_sale_days,
        "margin": margin,
        "alert": " · ".join(alerts) or "Monitorização",
        "unit": unit,
        "stockType": stock_type,
        "movementType": movement_type,
        "movementQuantity": movement_quantity,
        "warehouse": str(pick_value(normalized_row, ["armazem", "warehouse"]) or unit)[:80],
        "systemQuantity": system_quantity,
        "physicalQuantity": physical_quantity,
        "differenceQuantity": difference,
        "unitCost": round(unit_cost, 2),
        "stockValue": round(physical_quantity * unit_cost, 2),
        "location": str(pick_value(normalized_row, ["localizacao", "local", "prateleira"]) or "")[:100],
        "validationState": validation_state,
        "confidence": confidence,
    }


def row_to_debt(row: dict, normalized_row: dict, index: int) -> dict | None:
    debt_value = pick_value(normalized_row, DEBT_HEADERS)
    amount = parse_number(debt_value)
    entity = pick_value(normalized_row, ENTITY_HEADERS) or pick_value(normalized_row, ["nome"])
    if not entity or is_empty(debt_value) or amount == 0:
        return None

    due_days = int(abs(parse_number(pick_value(normalized_row, ["dias", "atraso", "prazo"]))))
    invoice = pick_value(normalized_row, ["documento", "fatura", "factura", "numero documento", "numero", "referencia"])
    issue_date = normalize_date(pick_value(normalized_row, ["data emissao", "emissao", "data doc", "data documento", "data"] ))
    due_date = normalize_date(pick_value(normalized_row, ["data vencimento", "vencimento", "due date", "duedate"] ))
    text = normalize(" ".join(str(value) for value in row.values()))
    return {
        "id": index + 9000,
        "invoice": str(invoice or f"DOC-{index + 1}")[:80],
        "entity": str(entity)[:120],
        "type": "Fornecedor" if "fornecedor" in text else "Cliente",
        "amount": abs(amount),
        "issueDate": issue_date,
        "dueDate": due_date,
        "dueDays": due_days,
        "state": "Em atraso" if due_days > 0 else "A vencer",
    }


def classify_movement(text: str, amount: float) -> tuple[str, str, str, int, str]:
    haystack = normalize(text)
    ranked = []
    for code, name, keywords, movement_type in ACCOUNT_RULES:
        hits = sum(1 for keyword in keywords if normalize(keyword) in haystack)
        if hits:
            ranked.append((hits, code, name, movement_type))
    ranked.sort(reverse=True)
    if ranked:
        hits, code, name, movement_type = ranked[0]
        return code, name, movement_type, min(96, 68 + hits * 12), "Classificação sugerida por palavras-chave SNC e contexto do movimento."
    if amount >= 0:
        return "78", "Outros rendimentos e ganhos", "Crédito", 52, "Entrada sem categoria específica identificada."
    return "68", "Outros gastos e perdas", "Débito", 52, "Saída sem categoria específica identificada."


def build_summary(source_name: str, rows_read: int, movements: list[dict]) -> dict:
    sales = sum(abs(item["amount"]) for item in movements if item["accountCode"] in {"71", "72", "78"})
    expenses = sum(abs(item["amount"]) for item in movements if item["accountCode"] in {"22", "24", "31", "61", "62", "63", "68"})
    profit = sales - expenses
    return {
        "sourceName": source_name,
        "rowsRead": rows_read,
        "sales": round(sales, 2),
        "expenses": round(expenses, 2),
        "profit": round(profit, 2),
        "margin": round((profit / sales) * 100, 1) if sales > 0 else 0,
    }


def build_issues(
    movements: list[dict],
    document_intelligence: dict | None = None,
    inventory: list[dict] | None = None,
    debts: list[dict] | None = None,
) -> list[dict]:
    """Create one actionable anomaly queue for the whole operational dataset."""
    issues: list[dict] = []
    seen: set[tuple[str, str]] = set()

    def add(document: str, source: str, value: float, issue: str, status: str) -> None:
        key = (normalize(document), normalize(issue))
        if key in seen or len(issues) >= 250:
            return
        seen.add(key)
        issues.append(
            {
                "id": 5000 + len(issues),
                "document": str(document or f"REG-{len(issues) + 1}")[:120],
                "source": str(source or "Não identificado")[:120],
                "value": f"{value:.2f} EUR",
                "issue": str(issue)[:180],
                "status": status,
            }
        )

    for movement in movements:
        if movement["confidence"] < 75:
            add(
                movement["description"],
                movement["entity"],
                abs(movement["amount"]),
                f"Classificação com baixa confiança ({movement['confidence']}%)",
                "Rever" if movement["confidence"] < 65 else "Classificar",
            )
        elif movement["accountCode"] == "68":
            add(
                movement["description"],
                movement["entity"],
                abs(movement["amount"]),
                "Conta genérica exige classificação específica",
                "Classificar",
            )

    for document in (document_intelligence or {}).get("documents", []):
        for validation in document.get("validations", []):
            validation_key = normalize(validation)
            if "duplicado" in validation_key:
                status = "Alerta"
            elif "nao corresponde" in validation_key or "invalida" in validation_key:
                status = "Rever"
            else:
                continue
            add(document["number"], document["entity"], abs(document["totalAmount"]), validation, status)

    for item in inventory or []:
        if item.get("differenceQuantity"):
            add(
                item["ref"],
                item.get("warehouse") or item.get("unit"),
                abs(item.get("differenceQuantity", 0) * item.get("unitCost", 0)),
                f"Divergência de inventário: sistema {item['systemQuantity']}, físico {item['physicalQuantity']}",
                "Alerta",
            )
        if item.get("physicalQuantity", 0) < 0:
            add(item["ref"], item.get("warehouse"), abs(item.get("stockValue", 0)), "Stock físico negativo", "Alerta")
        if item.get("confidence", 100) < 75:
            add(
                item["ref"],
                item.get("warehouse"),
                abs(item.get("stockValue", 0)),
                "Unidade ou tipo de stock não reconhecido",
                "Classificar",
            )

    for debt in debts or []:
        if debt.get("state") == "Em atraso":
            add(
                debt["invoice"],
                debt["entity"],
                abs(debt["amount"]),
                f"Pagamento vencido há {debt.get('dueDays', 0)} dias",
                "Alerta" if debt.get("dueDays", 0) > 30 else "Rever",
            )

    return issues


def pick_value(row: dict, candidates: list[str]):
    normalized_candidates = [normalize(candidate) for candidate in candidates]
    for normalized_candidate in normalized_candidates:
        for key, value in row.items():
            if str(key) == normalized_candidate and not is_empty(value):
                return value
    for normalized_candidate in normalized_candidates:
        for key, value in row.items():
            normalized_key = str(key)
            if normalized_candidate in normalized_key and not is_empty(value):
                return value
    return ""


def longest_text_value(row: dict) -> str:
    values = [str(value).strip() for value in row.values() if isinstance(value, str) and value.strip()]
    return max(values, key=len) if values else ""


def last_numeric_value(row: dict):
    values = [value for value in row.values() if parse_number(value) != 0]
    return values[-1] if values else ""


def parse_number(value) -> float:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value) if math.isfinite(float(value)) else 0
    text = str(value or "")
    negative = bool(re.search(r"\(\s*[\d,.]+", text) or re.search(r"-\s*\d", text))
    cleaned = re.sub(r"[^\d,.-]", "", text)
    cleaned = cleaned.replace("-", "")
    if not cleaned:
        return 0
    if "," in cleaned and "." in cleaned:
        decimal_separator = "," if cleaned.rfind(",") > cleaned.rfind(".") else "."
        thousands_separator = "." if decimal_separator == "," else ","
        cleaned = cleaned.replace(thousands_separator, "").replace(decimal_separator, ".")
    elif "," in cleaned:
        decimal_digits = len(cleaned.rsplit(",", 1)[-1])
        cleaned = cleaned.replace(",", "") if decimal_digits == 3 else cleaned.replace(",", ".")
    elif "." in cleaned:
        decimal_digits = len(cleaned.rsplit(".", 1)[-1])
        if cleaned.count(".") > 1 or decimal_digits == 3:
            cleaned = cleaned.replace(".", "")
    try:
        parsed = float(cleaned)
    except ValueError:
        return 0
    return -abs(parsed) if negative else parsed


def normalize_date(value) -> str:
    if not value:
        return "-"
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)[:24]


def is_empty(value) -> bool:
    return value is None or str(value).strip() == ""


def normalize(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value.lower())
    return "".join(char for char in normalized if unicodedata.category(char) != "Mn")


def dedupe_by(items: list[dict], key: str) -> list[dict]:
    deduped = {}
    for item in items:
        deduped[item[key]] = item
    return list(deduped.values())


def dedupe_debts(items: list[dict]) -> list[dict]:
    deduped = {}
    for item in items:
        deduped[f"{item['invoice']}-{item['entity']}-{item['type']}-{item['amount']}"] = item
    return list(deduped.values())
