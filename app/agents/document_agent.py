from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

SUPPORTED_SERIES = {"CUSA", "CNOV", "PUSA", "PNOV", "POFI"}
SUPPORTED_TYPES = {"FR", "FT", "NC"}

HEADER_ALIASES = {
    "id": {"id"},
    "documento": {"documento"},
    "data_doc": {"data doc", "data doc."},
    "entidade": {"entidade"},
    "total": {"total"},
    "total_liquido": {"total liquido", "total líquido"},
    "total_iva": {"total iva"},
    "estado": {"estado"},
    "doc_fornecedor": {"doc fornecedor", "doc. fornecedor"},
    "req_ext": {"nº enc / req ext", "nº enc. / req. ext.", "no enc / req ext"},
    "canal": {"canal de anuncios", "canal de anúncios"},
    "vendedor": {"vendedor"},
    "forma_liquidacao": {"f liquidacao", "f. liquidação", "forma de liquidacao", "forma de liquidação"},
}

DOC_RE = re.compile(r"^\s*(FR|FT|NC|GT|PF)\s+(CUSA|CNOV|PUSA|PNOV|POFI)/", re.IGNORECASE)


def _norm(value: Any) -> str:
    text = "" if value is None else str(value)
    text = "".join(
        char for char in unicodedata.normalize("NFKD", text.casefold())
        if not unicodedata.combining(char)
    )
    text = re.sub(r"[^a-z0-9º]+", " ", text)
    return " ".join(text.split())


def _money(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace("€", "").replace(" ", "")
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    elif "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid monetary value: {value!r}") from exc


def _date_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = "" if value is None else str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return value


@dataclass(frozen=True)
class BillingRecord:
    source_row: int
    id: str
    documento: str
    data_doc: Any
    entidade: str
    total: Decimal
    total_liquido: Decimal
    total_iva: Decimal
    estado: str
    doc_fornecedor: str
    req_ext: str
    canal: str
    vendedor: str
    forma_liquidacao: str
    doc_type: str
    serie: str
    special_kind: str | None = None

    @property
    def location(self) -> str:
        return "COIMBRA" if self.serie in {"CUSA", "CNOV"} else "PICOTO"

    @property
    def vehicle_group(self) -> str:
        if self.serie in {"CUSA", "PUSA"}:
            return "USADO"
        if self.serie in {"CNOV", "PNOV"}:
            return "NOVO"
        return "POFI / OFICINA"

    def as_row(self) -> list[Any]:
        return [
            self.id,
            self.documento,
            self.data_doc,
            self.entidade,
            float(self.total),
            float(self.total_liquido),
            float(self.total_iva),
            self.estado,
            self.doc_fornecedor,
            self.req_ext,
            self.canal,
            self.vendedor,
            self.forma_liquidacao,
        ]


class DocumentAgent:
    """Read and normalize a daily billing workbook without fixed column indexes."""

    def execute(self, content: bytes, filename: str = "upload.xlsx") -> dict:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        sheet, header_row, columns = self._find_source_table(workbook)
        records: list[BillingRecord] = []
        rejected: list[dict[str, Any]] = []

        for row_number, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            documento = self._cell(row, columns, "documento")
            if not documento:
                continue

            match = DOC_RE.match(str(documento))
            if not match:
                rejected.append({"row": row_number, "reason": "unsupported_document", "documento": str(documento)})
                continue

            doc_type, serie = match.group(1).upper(), match.group(2).upper()
            if doc_type not in SUPPORTED_TYPES or serie not in SUPPORTED_SERIES:
                rejected.append({"row": row_number, "reason": "outside_billing_model", "documento": str(documento)})
                continue

            estado = str(self._cell(row, columns, "estado") or "").strip()
            if "anulad" in _norm(estado):
                rejected.append({"row": row_number, "reason": "cancelled", "documento": str(documento)})
                continue

            total = _money(self._cell(row, columns, "total"))
            net = _money(self._cell(row, columns, "total_liquido"))
            vat = _money(self._cell(row, columns, "total_iva"))
            if doc_type == "NC":
                total, net, vat = -abs(total), -abs(net), -abs(vat)

            entidade = str(self._cell(row, columns, "entidade") or "").strip()
            canal = str(self._cell(row, columns, "canal") or "").strip()
            special_kind = self._special_kind(entidade, canal)

            records.append(
                BillingRecord(
                    source_row=row_number,
                    id=str(self._cell(row, columns, "id") or "").strip(),
                    documento=str(documento).strip(),
                    data_doc=_date_value(self._cell(row, columns, "data_doc")),
                    entidade=entidade,
                    total=total,
                    total_liquido=net,
                    total_iva=vat,
                    estado=estado,
                    doc_fornecedor=str(self._cell(row, columns, "doc_fornecedor") or "").strip(),
                    req_ext=str(self._cell(row, columns, "req_ext") or "").strip(),
                    canal=canal,
                    vendedor=str(self._cell(row, columns, "vendedor") or "").strip(),
                    forma_liquidacao=str(self._cell(row, columns, "forma_liquidacao") or "").strip(),
                    doc_type=doc_type,
                    serie=serie,
                    special_kind=special_kind,
                )
            )

        if not records:
            raise ValueError("No valid FR/FT/NC billing records found in workbook")

        return {
            "filename": filename,
            "sheet": sheet.title,
            "header_row": header_row,
            "records": records,
            "rejected": rejected,
        }

    def _find_source_table(self, workbook):
        required = {"documento", "total", "total_liquido", "total_iva", "estado", "vendedor"}
        for sheet in workbook.worksheets:
            for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
                normalized = [_norm(value) for value in values]
                columns: dict[str, int] = {}
                for logical, aliases in HEADER_ALIASES.items():
                    normalized_aliases = {_norm(alias) for alias in aliases}
                    for index, header in enumerate(normalized):
                        if header in normalized_aliases:
                            columns[logical] = index
                            break
                if required.issubset(columns):
                    return sheet, row_number, columns
        raise ValueError("Could not locate the billing header row by column names")

    @staticmethod
    def _cell(row: tuple[Any, ...], columns: dict[str, int], name: str) -> Any:
        index = columns.get(name)
        if index is None or index >= len(row):
            return None
        return row[index]

    @staticmethod
    def _special_kind(entidade: str, canal: str) -> str | None:
        entity = _norm(entidade)
        channel = _norm(canal)
        if "sucatas de ramil" in entity or re.search(r"\bsucata(s)?\b", entity):
            return "SUCATA"
        if re.search(r"\bsalvado(s)?\b", entity) or re.search(r"\bsalvado(s)?\b", channel):
            return "SALVADO"
        return None
