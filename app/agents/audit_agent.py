from __future__ import annotations

import re
from collections import defaultdict
from io import BytesIO

from openpyxl import load_workbook

from app.agents.document_agent import BillingRecord

REQUIRED_SHEETS = ("Faturação Separada", "Resumo Vendedores", "Mapa Diário")
DOC_RE = re.compile(r"^(FR|FT|NC)\s+(CUSA|CNOV|PUSA|PNOV|POFI)/", re.IGNORECASE)


class AuditAgent:
    """Autonomously open and validate the generated workbook against normalized source records."""

    def execute(self, source_records: list[BillingRecord], output_content: bytes) -> dict:
        workbook = load_workbook(BytesIO(output_content), data_only=False)
        checks: dict[str, bool] = {}
        errors: list[str] = []

        checks["three_required_sheets_present"] = all(name in workbook.sheetnames for name in REQUIRED_SHEETS)
        if not checks["three_required_sheets_present"]:
            errors.append("Missing one or more mandatory billing sheets")
            return self._result(checks, errors)

        separated = workbook["Faturação Separada"]
        summary = workbook["Resumo Vendedores"]
        daily = workbook["Mapa Diário"]

        expected_operational = [r for r in source_records if r.special_kind is None]
        output_records = self._read_separated_records(separated)

        checks["coimbra_picoto_separation"] = self._contains_labels(separated, {"COIMBRA", "PICOTO"})
        checks["novo_usado_separation"] = self._contains_labels(daily, {"USADO", "NOVO"})
        checks["sucatas_salvados_separation"] = self._contains_labels(daily, {"SUCATAS", "SALVADOS"})
        checks["fr_ft_separation"] = self._fr_ft_sections_are_distinct(separated)
        checks["nc_negative"] = all(
            record["total"] <= 0 and record["net"] <= 0 and record["vat"] <= 0
            for record in output_records if record["doc_type"] == "NC"
        )
        checks["nc_liquidado_pendente"] = self._contains_labels(daily, {"Liquidado", "Pendente"})
        checks["seller_summary_uses_net_total"] = self._check_seller_summary(summary, expected_operational)
        checks["daily_map_total_is_fr_plus_ft_plus_nc"] = self._check_map_formulas(daily)
        checks["workbook_totals_reconciled"] = self._records_reconcile(expected_operational, output_records)
        checks["source_reconciliation"] = checks["workbook_totals_reconciled"] and self._check_daily_components(daily, expected_operational)

        for name, passed in checks.items():
            if not passed:
                errors.append(name)

        return self._result(checks, errors)

    @staticmethod
    def _result(checks: dict[str, bool], errors: list[str]) -> dict:
        passed = bool(checks) and all(checks.values())
        return {
            "status": "COMPLETED" if passed else "FAILED",
            "valid": passed,
            "checks": checks,
            "errors": errors,
        }

    @staticmethod
    def _contains_labels(ws, labels: set[str]) -> bool:
        values = {str(cell.value).strip() for row in ws.iter_rows() for cell in row if cell.value is not None}
        return labels.issubset(values)

    @staticmethod
    def _read_separated_records(ws) -> list[dict]:
        records = []
        for row in ws.iter_rows(values_only=True):
            documento = row[1] if len(row) > 1 else None
            if not isinstance(documento, str):
                continue
            match = DOC_RE.match(documento.strip())
            if not match:
                continue
            try:
                total = float(row[4] or 0)
                net = float(row[5] or 0)
                vat = float(row[6] or 0)
            except (TypeError, ValueError):
                continue
            records.append({
                "id": str(row[0] or ""),
                "documento": documento.strip(),
                "doc_type": match.group(1).upper(),
                "serie": match.group(2).upper(),
                "total": total,
                "net": net,
                "vat": vat,
                "estado": str(row[7] or "").strip(),
                "vendedor": str(row[11] or "").strip(),
            })
        return records

    @staticmethod
    def _fr_ft_sections_are_distinct(ws) -> bool:
        labels = [str(ws.cell(row, 1).value or "") for row in range(1, ws.max_row + 1)]
        has_fr = any(label.startswith("FR ") for label in labels)
        has_ft = any(label.startswith("FT ") for label in labels)
        return has_fr and has_ft

    @staticmethod
    def _records_reconcile(expected: list[BillingRecord], actual: list[dict]) -> bool:
        expected_map = {
            (r.id, r.documento): (
                round(float(r.total), 2),
                round(float(r.total_liquido), 2),
                round(float(r.total_iva), 2),
                r.estado,
                r.vendedor,
            )
            for r in expected
        }
        actual_map = {
            (r["id"], r["documento"]): (
                round(r["total"], 2),
                round(r["net"], 2),
                round(r["vat"], 2),
                r["estado"],
                r["vendedor"],
            )
            for r in actual
        }
        return expected_map == actual_map

    @staticmethod
    def _check_seller_summary(ws, records: list[BillingRecord]) -> bool:
        expected = defaultdict(float)
        for r in records:
            state = r.estado if r.doc_type == "NC" else ""
            expected[(r.location, r.doc_type, state.casefold(), r.vendedor or "SEM VENDEDOR", r.serie)] += float(r.total_liquido)

        location = None
        doc_type = None
        state = ""
        series_by_col: dict[int, str] = {}
        seen = defaultdict(float)

        for row in ws.iter_rows():
            first = str(row[0].value or "").strip()
            if first in {"COIMBRA", "PICOTO"}:
                location = first
                doc_type = None
                state = ""
                continue
            if first in {"FR", "FT"} or first.startswith("NC —"):
                if first in {"FR", "FT"}:
                    doc_type, state = first, ""
                else:
                    doc_type = "NC"
                    state = "liquidado" if "LIQUIDAD" in first else "pendente"
                continue
            if first == "Vendedores":
                series_by_col = {}
                for cell in row[1:]:
                    value = str(cell.value or "")
                    match = re.search(r"(CUSA|CNOV|PUSA|PNOV|POFI)", value)
                    if match:
                        series_by_col[cell.column] = match.group(1)
                continue
            if not location or not doc_type or not series_by_col:
                continue
            if not first or first.startswith("TOTAL ") or first.startswith("Resumo "):
                continue
            for col, serie in series_by_col.items():
                value = ws.cell(row[0].row, col).value
                if isinstance(value, (int, float)):
                    seen[(location, doc_type, state, first, serie)] += float(value)

        expected_nonzero = {k: round(v, 2) for k, v in expected.items() if round(v, 2) != 0}
        seen_nonzero = {k: round(v, 2) for k, v in seen.items() if round(v, 2) != 0}
        return expected_nonzero == seen_nonzero

    @staticmethod
    def _check_map_formulas(ws) -> bool:
        component_rows = []
        for row in range(1, ws.max_row + 1):
            group = str(ws.cell(row, 1).value or "")
            serie = str(ws.cell(row, 2).value or "")
            if group in {"USADO", "NOVO", "POFI / OFICINA"} and serie in {"CUSA", "CNOV", "PUSA", "PNOV", "POFI"}:
                expected_i = f"=C{row}+E{row}+G{row}"
                expected_j = f"=D{row}+F{row}+H{row}"
                if ws.cell(row, 9).value != expected_i or ws.cell(row, 10).value != expected_j:
                    return False
                component_rows.append(row)
        return len(component_rows) == 5

    @staticmethod
    def _check_daily_components(ws, records: list[BillingRecord]) -> bool:
        expected = {}
        for serie in ("CUSA", "CNOV", "PUSA", "PNOV", "POFI"):
            subset = [r for r in records if r.serie == serie]
            values = []
            for doc_type in ("FR", "FT", "NC"):
                docs = [r for r in subset if r.doc_type == doc_type]
                values.extend([
                    round(sum(float(r.total) for r in docs), 2),
                    round(sum(float(r.total_liquido) for r in docs), 2),
                ])
            expected[serie] = values

        actual = {}
        for row in range(1, ws.max_row + 1):
            group = str(ws.cell(row, 1).value or "")
            serie = str(ws.cell(row, 2).value or "")
            if group in {"USADO", "NOVO", "POFI / OFICINA"} and serie in expected:
                actual[serie] = [round(float(ws.cell(row, col).value or 0), 2) for col in range(3, 9)]
        return actual == expected
