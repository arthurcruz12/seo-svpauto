from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.agents.document_agent import BillingRecord

HEADERS = [
    "ID", "Documento", "Data Doc.", "Entidade", "Total", "Total liquido", "Total IVA",
    "Estado", "Doc. Fornecedor", "Nº Enc. / Req. Ext.", "Canal de Anúncios", "Vendedor",
    "F. Liquidação",
]
SERIES_ORDER = ["CUSA", "CNOV", "PUSA", "PNOV", "POFI"]
DOC_ORDER = ["FR", "FT", "NC"]


def _money_sum(records: Iterable[BillingRecord], field: str) -> float:
    return float(sum((getattr(record, field) for record in records), start=0))


def _date_label(records: list[BillingRecord]) -> str:
    values = [r.data_doc for r in records if isinstance(r.data_doc, datetime)]
    if not values:
        return ""
    return values[0].strftime("%d/%m/%Y")


class BillingAgent:
    """Generate the mandatory three-sheet SVP Auto daily billing workbook."""

    def execute(self, document_result: dict) -> dict:
        records: list[BillingRecord] = list(document_result["records"])
        operational = [r for r in records if r.special_kind is None]
        specials = [r for r in records if r.special_kind is not None]

        workbook = Workbook()
        workbook.remove(workbook.active)

        _, subtotal_rows = self._build_separated(workbook, operational)
        self._build_sellers(workbook, operational)
        self._build_daily_map(workbook, operational, specials)

        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"

        buffer = BytesIO()
        workbook.save(buffer)
        return {
            "content": buffer.getvalue(),
            "sheet_names": workbook.sheetnames,
            "subtotal_rows": subtotal_rows,
            "records": operational,
            "special_records": specials,
        }

    def _build_separated(self, workbook: Workbook, records: list[BillingRecord]):
        ws = workbook.create_sheet("Faturação Separada")
        date_text = _date_label(records)
        ws.append([f"FATURAÇÃO SEPARADA — {date_text}"])
        ws.append(["GT, PF, documentos anulados e tipos fora de FR/FT/NC removidos • NC negativas • NC separadas por Liquidado / Pendente"])
        ws.append([])
        subtotal_rows: dict[tuple[str, str, str | None], int] = {}

        for location, series in (("COIMBRA", ["CUSA", "CNOV"]), ("PICOTO", ["PUSA", "PNOV", "POFI"])):
            ws.append([location])
            ws.cell(ws.max_row, 1).font = Font(bold=True)
            for doc_type in DOC_ORDER:
                for serie in series:
                    states = ["Liquidado", "Pendente"] if doc_type == "NC" else [None]
                    for state in states:
                        group = [
                            r for r in records
                            if r.serie == serie and r.doc_type == doc_type
                            and (state is None or r.estado.casefold() == state.casefold())
                        ]
                        if not group:
                            continue
                        label = f"{doc_type} {serie}" if state is None else f"{doc_type} {serie} — {state.upper()}S"
                        ws.append([label])
                        ws.cell(ws.max_row, 1).font = Font(bold=True)
                        ws.append(HEADERS)
                        header_row = ws.max_row
                        for cell in ws[header_row]:
                            cell.font = Font(bold=True)
                        start = ws.max_row + 1
                        for record in group:
                            ws.append(record.as_row())
                        end = ws.max_row
                        ws.append([
                            f"SUBTOTAL {label} ({len(group)})", None, None, None,
                            f"=SUM(E{start}:E{end})", f"=SUM(F{start}:F{end})", f"=SUM(G{start}:G{end})",
                        ])
                        subtotal_rows[(doc_type, serie, state)] = ws.max_row
                        ws.append([])

            location_rows = [
                row for (_, serie, _), row in subtotal_rows.items()
                if serie in series
            ]
            if location_rows:
                refs_e = ",".join(f"E{row}" for row in location_rows)
                refs_f = ",".join(f"F{row}" for row in location_rows)
                refs_g = ",".join(f"G{row}" for row in location_rows)
                ws.append([
                    f"TOTAL {location}", None, None, None,
                    f"=SUM({refs_e})", f"=SUM({refs_f})", f"=SUM({refs_g})",
                ])
                subtotal_rows[("TOTAL", location, None)] = ws.max_row
                ws.append([])

        coimbra_row = subtotal_rows.get(("TOTAL", "COIMBRA", None))
        picoto_row = subtotal_rows.get(("TOTAL", "PICOTO", None))
        if coimbra_row and picoto_row:
            ws.append([
                "TOTAL GERAL", None, None, None,
                f"=E{coimbra_row}+E{picoto_row}", f"=F{coimbra_row}+F{picoto_row}", f"=G{coimbra_row}+G{picoto_row}",
            ])

        self._format_detail_sheet(ws)
        return ws, subtotal_rows

    def _build_sellers(self, workbook: Workbook, records: list[BillingRecord]):
        ws = workbook.create_sheet("Resumo Vendedores")
        date_text = _date_label(records)
        ws.append([f"RESUMO POR VENDEDOR — {date_text}"])
        ws.append(["Total líquido agrupado por vendedor, tipo de documento, série e estado das NC"])
        ws.append([])

        for location, series in (("COIMBRA", ["CUSA", "CNOV"]), ("PICOTO", ["PUSA", "PNOV", "POFI"])):
            ws.append([location])
            ws.cell(ws.max_row, 1).font = Font(bold=True)
            for doc_type in DOC_ORDER:
                states = ["Liquidado", "Pendente"] if doc_type == "NC" else [None]
                for state in states:
                    filtered = [
                        r for r in records if r.serie in series and r.doc_type == doc_type
                        and (state is None or r.estado.casefold() == state.casefold())
                    ]
                    if not filtered:
                        continue
                    title = doc_type if state is None else f"NC — {state.upper()}S"
                    ws.append([title])
                    ws.cell(ws.max_row, 1).font = Font(bold=True)
                    header = ["Vendedores", *[f"Total líquido {serie}" for serie in series]]
                    ws.append(header)
                    for cell in ws[ws.max_row]:
                        cell.font = Font(bold=True)

                    vendors = sorted({r.vendedor or "SEM VENDEDOR" for r in filtered})
                    start = ws.max_row + 1
                    for vendor in vendors:
                        row = [vendor]
                        for serie in series:
                            amount = sum(
                                (r.total_liquido for r in filtered if (r.vendedor or "SEM VENDEDOR") == vendor and r.serie == serie),
                                start=0,
                            )
                            row.append(float(amount) if amount else None)
                        ws.append(row)
                    end = ws.max_row
                    total_row = [f"TOTAL {title}"]
                    for col in range(2, 2 + len(series)):
                        letter = get_column_letter(col)
                        total_row.append(f"=SUM({letter}{start}:{letter}{end})")
                    ws.append(total_row)
                    ws.append([])

        ws.append(["Resumo calculado exclusivamente a partir de Total líquido. NC separadas por Liquidado e Pendente."])
        self._format_summary_sheet(ws)

    def _build_daily_map(self, workbook: Workbook, records: list[BillingRecord], specials: list[BillingRecord]):
        ws = workbook.create_sheet("Mapa Diário")
        date_text = _date_label(records + specials)
        ws.append(["SVP AUTO — RESUMO DIÁRIO DE FATURAÇÃO"])
        ws.append([f"{date_text} • Coimbra e Picoto separados • Usado/Novo separados • FR + FT + NC • detalhe NC por estado"])
        ws.append([])

        component_rows: dict[str, int] = {}
        for location, pairs in (
            ("COIMBRA", [("USADO", "CUSA"), ("NOVO", "CNOV")]),
            ("PICOTO", [("USADO", "PUSA"), ("NOVO", "PNOV")]),
        ):
            ws.append([location])
            ws.cell(ws.max_row, 1).font = Font(bold=True)
            ws.append(["Grupo", "Série", "FR c/ IVA", "FR s/ IVA", "FT c/ IVA", "FT s/ IVA", "NC c/ IVA", "NC s/ IVA", "TOTAL c/ IVA", "TOTAL s/ IVA"])
            for group, serie in pairs:
                values = self._series_components(records, serie)
                row_number = ws.max_row + 1
                ws.append([group, serie, *values, f"=C{row_number}+E{row_number}+G{row_number}", f"=D{row_number}+F{row_number}+H{row_number}"])
                component_rows[serie] = ws.max_row
            a, b = [component_rows[serie] for _, serie in pairs]
            ws.append([
                f"TOTAL {location} — USADO + NOVO", None,
                f"=C{a}+C{b}", f"=D{a}+D{b}", f"=E{a}+E{b}", f"=F{a}+F{b}",
                f"=G{a}+G{b}", f"=H{a}+H{b}", f"=I{a}+I{b}", f"=J{a}+J{b}",
            ])
            component_rows[f"TOTAL_{location}"] = ws.max_row
            ws.append([])

        ws.append(["Grupo", "Série", "FR c/ IVA", "FR s/ IVA", "FT c/ IVA", "FT s/ IVA", "NC c/ IVA", "NC s/ IVA", "TOTAL c/ IVA", "TOTAL s/ IVA"])
        pofi = self._series_components(records, "POFI")
        pofi_row = ws.max_row + 1
        ws.append(["POFI / OFICINA", "POFI", *pofi, f"=C{pofi_row}+E{pofi_row}+G{pofi_row}", f"=D{pofi_row}+F{pofi_row}+H{pofi_row}"])
        component_rows["POFI"] = ws.max_row
        picoto = component_rows["TOTAL_PICOTO"]
        pofi_row = component_rows["POFI"]
        ws.append([
            "TOTAL PICOTO GERAL — USADO + NOVO + POFI", None,
            f"=C{picoto}+C{pofi_row}", f"=D{picoto}+D{pofi_row}", f"=E{picoto}+E{pofi_row}", f"=F{picoto}+F{pofi_row}",
            f"=G{picoto}+G{pofi_row}", f"=H{picoto}+H{pofi_row}", f"=I{picoto}+I{pofi_row}", f"=J{picoto}+J{pofi_row}",
        ])
        component_rows["TOTAL_PICOTO_GERAL"] = ws.max_row
        ws.append([])

        coimbra = component_rows["TOTAL_COIMBRA"]
        picoto_general = component_rows["TOTAL_PICOTO_GERAL"]
        ws.append([
            "TOTAL GERAL SVP AUTO", None,
            f"=C{coimbra}+C{picoto_general}", f"=D{coimbra}+D{picoto_general}", f"=E{coimbra}+E{picoto_general}", f"=F{coimbra}+F{picoto_general}",
            f"=G{coimbra}+G{picoto_general}", f"=H{coimbra}+H{picoto_general}", f"=I{coimbra}+I{picoto_general}", f"=J{coimbra}+J{picoto_general}",
        ])
        ws.append([])
        ws.append([])

        ws.append(["DETALHE DAS NOTAS DE CRÉDITO — POR ESTADO"])
        ws.append(["Unidade", "Série", "Estado", "NC c/ IVA", "NC s/ IVA"])
        for serie in SERIES_ORDER:
            location = "COIMBRA" if serie in {"CUSA", "CNOV"} else "PICOTO"
            for state in ("Liquidado", "Pendente"):
                nc = [r for r in records if r.serie == serie and r.doc_type == "NC" and r.estado.casefold() == state.casefold()]
                ws.append([location, serie, state, _money_sum(nc, "total"), _money_sum(nc, "total_liquido")])

        ws.append([])
        ws.append([])
        self._append_special_block(ws, "SUCATAS", "SUCATA", specials)
        ws.append([])
        ws.append([])
        self._append_special_block(ws, "SALVADOS", "SALVADO", specials)
        ws.append([])
        ws.append([])
        ws.append(["Leitura: Coimbra = CUSA + CNOV. Picoto = PUSA + PNOV; POFI separado. NC negativas. Sucatas e Salvados em blocos independentes."])
        self._format_map_sheet(ws)

    @staticmethod
    def _series_components(records: list[BillingRecord], serie: str) -> list[float]:
        subset = [r for r in records if r.serie == serie]
        values = []
        for doc_type in ("FR", "FT", "NC"):
            docs = [r for r in subset if r.doc_type == doc_type]
            values.extend([_money_sum(docs, "total"), _money_sum(docs, "total_liquido")])
        return values

    @staticmethod
    def _append_special_block(ws, title: str, kind: str, specials: list[BillingRecord]):
        ws.append([title])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
        ws.append(["Unidade", "Com IVA", "Sem IVA", "Observação", "Estado"])
        start = ws.max_row + 1
        for location in ("COIMBRA", "PICOTO"):
            rows = [r for r in specials if r.special_kind == kind and r.location == location]
            states = ", ".join(sorted({r.estado for r in rows if r.estado})) or "—"
            observation = f"{len(rows)} movimento(s) explicitamente classificado(s)" if rows else "Sem classificação explícita"
            ws.append([location, _money_sum(rows, "total"), _money_sum(rows, "total_liquido"), observation, states])
        end = ws.max_row
        ws.append([f"TOTAL {title}", f"=SUM(B{start}:B{end})", f"=SUM(C{start}:C{end})"])

    @staticmethod
    def _format_detail_sheet(ws):
        ws.freeze_panes = "A4"
        widths = [16, 22, 14, 42, 14, 14, 14, 14, 16, 20, 28, 28, 16]
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top")
        for col in ("E", "F", "G"):
            for cell in ws[col]:
                if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                    cell.number_format = '#,##0.00 [$€-pt-PT]'

    @staticmethod
    def _format_summary_sheet(ws):
        ws.freeze_panes = "A4"
        ws.column_dimensions["A"].width = 34
        for col in range(2, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20
        for row in ws.iter_rows():
            for cell in row:
                if cell.column > 1 and (isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("="))):
                    cell.number_format = '#,##0.00 [$€-pt-PT]'

    @staticmethod
    def _format_map_sheet(ws):
        ws.freeze_panes = "A4"
        ws.column_dimensions["A"].width = 42
        for col in range(2, 11):
            ws.column_dimensions[get_column_letter(col)].width = 18
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top")
                if cell.column >= 2 and (isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("="))):
                    cell.number_format = '#,##0.00 [$€-pt-PT]'
