from __future__ import annotations

from datetime import date, datetime
from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Visual language from the approved SVP Auto billing model.
NAVY = "1F4E78"
BLUE = "4472C4"
LIGHT_BLUE = "DDEBF7"
GREEN = "E2F0D9"
COIMBRA_BLUE = "2F75B5"
PICOTO_GOLD = "997300"
PICOTO_LIGHT = "FFF2CC"
LIGHT_GREY = "E7E6E6"
RED = "C00000"
RED_LIGHT = "FCE4D6"
PURPLE = "8064A2"
PURPLE_LIGHT = "E4DFEC"
TEAL = "0097A7"
WHITE = "FFFFFF"
DARK = "1F1F1F"
GREY = "666666"
MONEY_FORMAT = '#,##0.00 "€";-#,##0.00 "€"'


def _fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)


def _paint_row(
    ws: Any,
    row_number: int,
    *,
    fill_color: str | None = None,
    font_color: str = DARK,
    bold: bool = True,
    italic: bool = False,
    horizontal: str | None = None,
    size: int = 10,
    max_col: int | None = None,
) -> None:
    limit = max_col or ws.max_column
    for cell in ws[row_number][:limit]:
        if fill_color:
            cell.fill = _fill(fill_color)
        cell.font = Font(
            name="Calibri",
            size=size,
            bold=bold,
            italic=italic,
            color=font_color,
        )
        cell.alignment = Alignment(
            vertical="center",
            horizontal=horizontal,
            wrap_text=True,
        )


def _set_default_alignment(ws: Any) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)


def _set_money_format(ws: Any, columns: tuple[str, ...]) -> None:
    for column in columns:
        for cell in ws[column]:
            if isinstance(cell.value, (int, float)) or (
                isinstance(cell.value, str) and cell.value.startswith("=")
            ):
                cell.number_format = MONEY_FORMAT


def _format_detail_sheet(ws: Any) -> None:
    ws.freeze_panes = "A4"
    widths = [16, 22, 14, 42, 14, 14, 14, 14, 16, 20, 28, 28, 16]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    _set_default_alignment(ws)
    _set_money_format(ws, ("E", "F", "G"))

    for cell in ws["C"]:
        if isinstance(cell.value, (datetime, date)):
            cell.number_format = "dd/mm/yyyy"

    for row_number in range(1, ws.max_row + 1):
        label = str(ws.cell(row_number, 1).value or "").strip()
        document = str(ws.cell(row_number, 2).value or "").strip().upper()

        if row_number == 1:
            _paint_row(ws, row_number, fill_color=NAVY, font_color=WHITE, bold=True, horizontal="center", size=11, max_col=13)
        elif row_number == 2:
            _paint_row(ws, row_number, font_color=GREY, bold=False, italic=True, max_col=13)
        elif label in {"COIMBRA", "PICOTO"}:
            _paint_row(ws, row_number, fill_color=NAVY, font_color=WHITE, bold=True, max_col=13)
        elif label == "ID":
            _paint_row(ws, row_number, fill_color=BLUE, font_color=WHITE, bold=True, horizontal="center", max_col=13)
        elif label.startswith(("FR ", "FT ", "NC ")):
            _paint_row(
                ws,
                row_number,
                fill_color=LIGHT_BLUE,
                font_color=RED if label.startswith("NC ") else NAVY,
                bold=True,
                horizontal="center",
                max_col=13,
            )
        elif label.startswith("SUBTOTAL"):
            _paint_row(
                ws,
                row_number,
                fill_color=GREEN,
                font_color=RED if " NC " in f" {label.upper()} " else NAVY,
                bold=True,
                max_col=13,
            )
        elif label == "TOTAL GERAL":
            _paint_row(ws, row_number, fill_color=NAVY, font_color=WHITE, bold=True, max_col=13)
        elif label.startswith("TOTAL "):
            _paint_row(ws, row_number, fill_color=LIGHT_BLUE, font_color=NAVY, bold=True, max_col=13)

        if document.startswith("NC "):
            for column in (5, 6, 7):
                ws.cell(row_number, column).font = Font(name="Calibri", size=10, color=RED)


def _format_summary_sheet(ws: Any) -> None:
    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 34
    for column in range(2, 8):
        ws.column_dimensions[get_column_letter(column)].width = 20

    _set_default_alignment(ws)
    for column in range(2, 8):
        _set_money_format(ws, (get_column_letter(column),))

    for row_number in range(1, ws.max_row + 1):
        label = str(ws.cell(row_number, 1).value or "").strip()
        upper = label.upper()

        if row_number == 1:
            _paint_row(ws, row_number, fill_color=NAVY, font_color=WHITE, bold=True, horizontal="center", size=11, max_col=6)
        elif row_number == 2:
            _paint_row(ws, row_number, font_color=GREY, bold=False, italic=True, max_col=6)
        elif label in {"COIMBRA", "PICOTO"}:
            _paint_row(ws, row_number, fill_color=NAVY, font_color=WHITE, bold=True, max_col=6)
        elif label == "Vendedores":
            _paint_row(ws, row_number, fill_color=BLUE, font_color=WHITE, bold=True, horizontal="center", max_col=6)
        elif upper in {"FR", "FT"} or upper.startswith("NC"):
            _paint_row(
                ws,
                row_number,
                fill_color=LIGHT_BLUE,
                font_color=RED if upper.startswith("NC") else NAVY,
                bold=True,
                max_col=6,
            )
        elif upper.startswith("TOTAL "):
            _paint_row(
                ws,
                row_number,
                fill_color=GREEN,
                font_color=RED if "NC" in upper else NAVY,
                bold=True,
                max_col=6,
            )


def _format_map_sheet(ws: Any) -> None:
    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 42
    for column in range(2, 11):
        ws.column_dimensions[get_column_letter(column)].width = 18

    _set_default_alignment(ws)
    for column in range(2, 11):
        _set_money_format(ws, (get_column_letter(column),))

    current_section: str | None = None
    special_section: str | None = None
    in_nc_detail = False

    for row_number in range(1, ws.max_row + 1):
        label = str(ws.cell(row_number, 1).value or "").strip()
        upper = label.upper()
        state = str(ws.cell(row_number, 3).value or "").strip().casefold()

        if row_number == 1:
            _paint_row(ws, row_number, fill_color=NAVY, font_color=WHITE, bold=True, horizontal="center", size=11, max_col=10)
            continue
        if row_number == 2:
            _paint_row(ws, row_number, font_color=GREY, bold=False, italic=True, max_col=10)
            continue

        if label == "COIMBRA":
            current_section = "COIMBRA"
            special_section = None
            in_nc_detail = False
            _paint_row(ws, row_number, fill_color=COIMBRA_BLUE, font_color=WHITE, bold=True, max_col=10)
            continue
        if label == "PICOTO":
            current_section = "PICOTO"
            special_section = None
            in_nc_detail = False
            _paint_row(ws, row_number, fill_color=PICOTO_GOLD, font_color=WHITE, bold=True, max_col=10)
            continue

        if upper == "DETALHE DAS NOTAS DE CRÉDITO — POR ESTADO":
            in_nc_detail = True
            special_section = None
            _paint_row(ws, row_number, fill_color=RED, font_color=WHITE, bold=True, max_col=10)
            continue
        if upper == "SUCATAS":
            in_nc_detail = False
            special_section = "SUCATAS"
            _paint_row(ws, row_number, fill_color=PURPLE, font_color=WHITE, bold=True, max_col=10)
            continue
        if upper == "SALVADOS":
            in_nc_detail = False
            special_section = "SALVADOS"
            _paint_row(ws, row_number, fill_color=TEAL, font_color=WHITE, bold=True, max_col=10)
            continue

        if upper == "TOTAL GERAL SVP AUTO":
            _paint_row(ws, row_number, fill_color=NAVY, font_color=WHITE, bold=True, max_col=10)
        elif upper.startswith("TOTAL PICOTO GERAL"):
            _paint_row(ws, row_number, fill_color=LIGHT_GREY, font_color=NAVY, bold=True, max_col=10)
        elif upper.startswith("TOTAL COIMBRA"):
            _paint_row(ws, row_number, fill_color=LIGHT_BLUE, font_color=NAVY, bold=True, max_col=10)
        elif upper.startswith("TOTAL PICOTO"):
            _paint_row(ws, row_number, fill_color=PICOTO_LIGHT, font_color=DARK, bold=True, max_col=10)
        elif upper == "TOTAL SUCATAS":
            _paint_row(ws, row_number, fill_color=PURPLE_LIGHT, font_color=NAVY, bold=True, max_col=10)
        elif upper == "TOTAL SALVADOS":
            _paint_row(ws, row_number, fill_color=LIGHT_BLUE, font_color=NAVY, bold=True, max_col=10)
        elif label == "Grupo":
            header_fill = COIMBRA_BLUE if current_section == "COIMBRA" else PICOTO_GOLD if current_section == "PICOTO" else BLUE
            _paint_row(ws, row_number, fill_color=header_fill, font_color=WHITE, bold=True, horizontal="center", max_col=10)
        elif label == "Unidade":
            if in_nc_detail:
                _paint_row(ws, row_number, fill_color=RED_LIGHT, font_color=RED, bold=True, horizontal="center", max_col=10)
            elif special_section == "SUCATAS":
                _paint_row(ws, row_number, fill_color=PURPLE_LIGHT, font_color=NAVY, bold=True, horizontal="center", max_col=10)
            elif special_section == "SALVADOS":
                _paint_row(ws, row_number, fill_color=LIGHT_BLUE, font_color=NAVY, bold=True, horizontal="center", max_col=10)
        elif upper == "POFI / OFICINA":
            _paint_row(ws, row_number, fill_color=LIGHT_GREY, font_color=NAVY, bold=True, max_col=10)
        elif upper.startswith("LEITURA:"):
            _paint_row(ws, row_number, font_color=GREY, bold=False, italic=True, max_col=10)

        if in_nc_detail and state in {"liquidado", "pendente"}:
            if state == "pendente":
                ws.cell(row_number, 3).font = Font(name="Calibri", size=10, bold=True, color=RED)
            for column in (4, 5):
                ws.cell(row_number, column).font = Font(name="Calibri", size=10, color=RED)


def install_billing_styles(billing_agent_class: type[Any]) -> None:
    """Attach the approved SVP visual model without changing billing calculations."""
    billing_agent_class._format_detail_sheet = staticmethod(_format_detail_sheet)
    billing_agent_class._format_summary_sheet = staticmethod(_format_summary_sheet)
    billing_agent_class._format_map_sheet = staticmethod(_format_map_sheet)
