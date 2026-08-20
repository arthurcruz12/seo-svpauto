from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.agents.audit_agent import AuditAgent
from app.agents.billing_agent import BillingAgent
from app.agents.document_agent import DocumentAgent
from app.agents.manager import AgentManager


HEADERS = [
    "ID", "Documento", "Data Doc.", "Entidade", "Total", "Total liquido", "Total IVA",
    "Estado", "Doc. Fornecedor", "Nº Enc. / Req. Ext.", "Canal de Anúncios", "Vendedor",
    "F. Liquidação",
]


def _source_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    ws.append(["1", "FR CUSA/1", "2026-08-20", "Cliente A", 123, 100, 23, "Liquidado", "", "", "Balcão", "1005 - VENDEDOR A", "MB"])
    ws.append(["2", "FT CUSA/1", "2026-08-20", "Cliente B", 61.5, 50, 11.5, "Pendente", "", "", "Site", "1018 - VENDEDOR B", "TR"])
    ws.append(["3", "NC CNOV/1", "2026-08-20", "Cliente C", 12.3, 10, 2.3, "Liquidado", "", "", "Balcão", "1005 - VENDEDOR A", "NU"])
    ws.append(["4", "FR PUSA/1", "2026-08-20", "Cliente D", 246, 200, 46, "Liquidado", "", "", "Balcão", "2004 - VENDEDOR C", "MB"])
    ws.append(["5", "FT PNOV/1", "2026-08-20", "Cliente E", 123, 100, 23, "Pendente", "", "", "Balcão", "2019 - VENDEDOR D", "MB"])
    ws.append(["6", "NC PUSA/1", "2026-08-20", "Cliente F", 24.6, 20, 4.6, "Pendente", "", "", "Balcão", "2004 - VENDEDOR C", "TB"])
    ws.append(["7", "FR POFI/1", "2026-08-20", "Cliente Oficina", 61.5, 50, 11.5, "Liquidado", "", "", "Balcão", "2106 - OFICINA", "MB"])
    ws.append(["8", "FT PUSA/2", "2026-08-20", "SUCATAS DE RAMIL, S.A", 300, 300, 0, "Pendente", "", "", "Balcão", "2000 - SVP", "TB"])
    ws.append(["9", "GT CUSA/1", "2026-08-20", "Ignorar", 999, 999, 0, "Pendente", "", "", "Balcão", "1005 - VENDEDOR A", "TB"])
    ws.append(["10", "FR CUSA/2", "2026-08-20", "Anulado", 999, 999, 0, "Anulado", "", "", "Balcão", "1005 - VENDEDOR A", "TB"])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_document_agent_normalizes_credit_notes_and_excludes_outside_model():
    result = DocumentAgent().execute(_source_workbook(), "billing.xlsx")
    by_document = {record.documento: record for record in result["records"]}

    assert "GT CUSA/1" not in by_document
    assert "FR CUSA/2" not in by_document
    assert by_document["NC CNOV/1"].total < 0
    assert by_document["NC CNOV/1"].total_liquido < 0
    assert by_document["FT PUSA/2"].special_kind == "SUCATA"


def test_billing_agent_generates_exact_three_sheets_and_autonomous_audit_passes():
    document_result = DocumentAgent().execute(_source_workbook(), "billing.xlsx")
    result = BillingAgent().execute(document_result)

    assert result["sheet_names"] == ["Faturação Separada", "Resumo Vendedores", "Mapa Diário"]
    audit = AuditAgent().execute(document_result["records"], result["content"])
    assert audit["status"] == "COMPLETED"
    assert audit["valid"] is True
    assert all(audit["checks"].values())


def test_auditor_fails_when_generated_workbook_is_tampered():
    document_result = DocumentAgent().execute(_source_workbook(), "billing.xlsx")
    result = BillingAgent().execute(document_result)

    wb = load_workbook(BytesIO(result["content"]))
    ws = wb["Faturação Separada"]
    for row in range(1, ws.max_row + 1):
        document = ws.cell(row, 2).value
        if isinstance(document, str) and document.startswith(("FR ", "FT ", "NC ")):
            ws.cell(row, 5).value = float(ws.cell(row, 5).value) + 1
            break

    buffer = BytesIO()
    wb.save(buffer)
    audit = AuditAgent().execute(document_result["records"], buffer.getvalue())

    assert audit["status"] == "FAILED"
    assert audit["valid"] is False
    assert audit["checks"]["workbook_totals_reconciled"] is False


def test_agent_manager_executes_real_billing_chain():
    result = AgentManager().execute_billing(_source_workbook(), "billing.xlsx")

    assert result["status"] == "COMPLETED"
    assert result["agents_used"] == ["DocumentAgent", "BillingAgent", "AuditAgent"]
    assert result["output_filename"].endswith(" - Separada, Resumo e Mapa Diário.xlsx")
    assert result["output_content"]
    assert result["audit"]["valid"] is True
