from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.agents import BillingAgent, BillingRecord
from app.billing_style import MONEY_FORMAT


def _record(
    source_row: int,
    documento: str,
    *,
    total: str = "123.00",
    liquid: str = "100.00",
    vat: str = "23.00",
    estado: str = "Liquidado",
    vendedor: str = "1005 - VENDEDOR TESTE",
    special_kind: str | None = None,
) -> BillingRecord:
    doc_type, series_number = documento.split(" ", 1)
    serie = series_number.split("/", 1)[0]
    return BillingRecord(
        source_row=source_row,
        id=str(source_row),
        documento=documento,
        data_doc=datetime(2026, 8, 20),
        entidade=f"Entidade {source_row}",
        total=Decimal(total),
        total_liquido=Decimal(liquid),
        total_iva=Decimal(vat),
        estado=estado,
        doc_fornecedor="",
        req_ext="",
        canal="Balcão",
        vendedor=vendedor,
        forma_liquidacao="MB",
        doc_type=doc_type,
        serie=serie,
        special_kind=special_kind,
    )


def _row_for(ws, label: str) -> int:
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 1).value or "").strip() == label:
            return row
    raise AssertionError(f"Label not found: {label}")


def _rgb(cell) -> str:
    value = cell.fill.fgColor.rgb or ""
    return value[-6:].upper()


def _font_rgb(cell) -> str:
    color = cell.font.color
    if not color or color.type != "rgb" or not color.rgb:
        return ""
    return color.rgb[-6:].upper()


def test_billing_agent_uses_approved_svp_color_model():
    records = [
        _record(1, "FR CUSA/1"),
        _record(2, "FT CNOV/1", estado="Pendente"),
        _record(3, "NC CUSA/1", total="-12.30", liquid="-10.00", vat="-2.30"),
        _record(4, "NC CNOV/1", total="-24.60", liquid="-20.00", vat="-4.60", estado="Pendente"),
        _record(5, "FR PUSA/1", vendedor="2004 - VENDEDOR PICOTO"),
        _record(6, "FT PNOV/1", estado="Pendente", vendedor="2019 - VENDEDOR PICOTO"),
        _record(7, "FR POFI/1", vendedor="2106 - OFICINA"),
        _record(8, "FT PUSA/2", special_kind="SUCATA", vendedor="2000 - SVP"),
        _record(9, "FT PUSA/3", special_kind="SALVADO", vendedor="2000 - SVP"),
    ]

    result = BillingAgent().execute({"records": records})
    workbook = load_workbook(BytesIO(result["content"]), data_only=False)

    separated = workbook["Faturação Separada"]
    assert _rgb(separated["A1"]) == "1F4E78"
    assert _font_rgb(separated["A1"]) == "FFFFFF"
    assert _rgb(separated.cell(_row_for(separated, "ID"), 1)) == "4472C4"
    assert _rgb(separated.cell(_row_for(separated, "SUBTOTAL NC CUSA — LIQUIDADOS (1)"), 1)) == "E2F0D9"
    assert _font_rgb(separated.cell(_row_for(separated, "SUBTOTAL NC CUSA — LIQUIDADOS (1)"), 1)) == "C00000"
    first_money = next(cell for cell in separated["E"] if isinstance(cell.value, (int, float)))
    assert first_money.number_format == MONEY_FORMAT

    sellers = workbook["Resumo Vendedores"]
    assert _rgb(sellers.cell(_row_for(sellers, "COIMBRA"), 1)) == "1F4E78"
    assert _rgb(sellers.cell(_row_for(sellers, "Vendedores"), 1)) == "4472C4"
    nc_total = _row_for(sellers, "TOTAL NC — LIQUIDADOS")
    assert _rgb(sellers.cell(nc_total, 1)) == "E2F0D9"
    assert _font_rgb(sellers.cell(nc_total, 1)) == "C00000"

    daily = workbook["Mapa Diário"]
    assert _rgb(daily.cell(_row_for(daily, "COIMBRA"), 1)) == "2F75B5"
    assert _rgb(daily.cell(_row_for(daily, "PICOTO"), 1)) == "997300"
    assert _rgb(daily.cell(_row_for(daily, "TOTAL GERAL SVP AUTO"), 1)) == "1F4E78"
    assert _rgb(daily.cell(_row_for(daily, "DETALHE DAS NOTAS DE CRÉDITO — POR ESTADO"), 1)) == "C00000"
    assert _rgb(daily.cell(_row_for(daily, "SUCATAS"), 1)) == "8064A2"
    assert _rgb(daily.cell(_row_for(daily, "SALVADOS"), 1)) == "0097A7"
